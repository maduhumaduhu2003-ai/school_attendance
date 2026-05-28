# ===============================
# Django core imports
# ===============================
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.contrib.auth import (
    authenticate,
    login,
    logout,
    update_session_auth_hash,
    get_user_model,
)
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.hashers import make_password
from django.contrib.auth.forms import PasswordChangeForm
from django.core.mail import send_mail
from django.core.paginator import Paginator
from django.conf import settings
from django.views.decorators.cache import never_cache
from django.db import IntegrityError, transaction
from django.db.models import Q, Prefetch
from django.db.models.deletion import ProtectedError
from django.utils import timezone
from django.utils.timezone import now
from django.core.exceptions import ValidationError
from django.core.files.storage import default_storage
from django.template.loader import render_to_string

# ===============================
# Python standard library
# ===============================
import csv
import random
import time
import re
import os
import logging
from datetime import date, datetime
from math import radians, cos, sin, sqrt, atan2

# ===============================
# Third-party libraries
# ===============================
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl.styles import Font, Alignment
import africastalking
import pandas as pd
from xhtml2pdf import pisa

# ===============================
# Local app imports
# ===============================
from .models import (
    User,
    TeacherProfile,
    StudentProfile,
    ParentProfile,
    Classroom,
    AcademicYear,
    Attendance,
    SMSLog,
    Stream,
    SchoolSettings,
    Enrollment,
)
from .forms import (
    SchoolSettingsForm,
    UserUpdateForm,
    AcademicYearForm,
)
from .utils import auto_lock_expired_academic_year, send_sms
from datetime import datetime
from django.views.decorators.csrf import csrf_protect, ensure_csrf_cookie
from django.http import JsonResponse
from django.urls import reverse


# Initialize logger
logger = logging.getLogger(__name__)

# Constants
SCHOOL_LAT = -6.92673
SCHOOL_LNG = 37.56749
MAX_DISTANCE_METERS = 500000
DEFAULT_PASSWORD = "Teacher@123"


def distance_in_meters(lat1, lon1, lat2, lon2):
    """Calculate distance between two GPS points (meters)"""
    try:
        R = 6371000
        dlat = radians(lat2 - lat1)
        dlon = radians(lon2 - lon1)

        a = sin(dlat / 2) ** 2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon / 2) ** 2
        c = 2 * atan2(sqrt(a), sqrt(1 - a))

        return R * c
    except Exception:
        return None


# ================= HELPER FUNCTIONS =================
# used to fetch teacher class and stream for current active accademic year
def get_teacher_current_assignment(teacher):
    active_year = AcademicYear.objects.filter(is_active=True).first()
    if not active_year:
        return None, None
    
    enrollment = Enrollment.objects.filter(
        class_teacher=teacher,
        academic_year=active_year
    ).select_related('classroom', 'stream').first()
    
    if enrollment:
        return enrollment.classroom, enrollment.stream
    return None, None


def get_student_current_enrollment(student, academic_year=None):
    """Get student's current enrollment"""
    if not academic_year:
        academic_year = AcademicYear.objects.filter(is_active=True).first()
    
    return student.enrollments.filter(academic_year=academic_year).first()


def get_students_by_teacher_scope(classroom, stream=None, academic_year=None):
    """Return students linked to the class/stream via Enrollment."""
    qs = StudentProfile.objects.filter(
        enrollments__classroom=classroom,
        enrollments__status='Active'
    )
    if stream:
        qs = qs.filter(enrollments__stream=stream)
    if academic_year:
        qs = qs.filter(enrollments__academic_year=academic_year)
    return qs.distinct()


# ================= AUTHENTICATION VIEWS =================
@never_cache
@ensure_csrf_cookie
def login_view(request):
    admin_exists = User.objects.filter(role='admin').exists()

    if request.method == 'POST':
        # Get credentials
        username = request.POST.get('username')
        password = request.POST.get('password')
        lat = request.POST.get('lat')
        lng = request.POST.get('lng')

        # Validate location
        try:
            lat = float(lat) if lat else None
            lng = float(lng) if lng else None
        except (TypeError, ValueError):
            messages.error(request, "Location error: allow GPS access.")
            return redirect('login')

        # Check distance if location provided
        if lat and lng:
            distance = distance_in_meters(lat, lng, SCHOOL_LAT, SCHOOL_LNG)
            if distance is None:
                messages.error(request, "Could not calculate distance.")
                return redirect('login')

            if distance > MAX_DISTANCE_METERS:
                messages.error(request, "Access denied: you are outside school area.")
                return redirect('login')

        # Authenticate user
        user = authenticate(request, username=username, password=password)

        if not user:
            messages.error(request, "Invalid email or password!")
            return redirect('login')

        # ========== CHECK TEACHER STATUS (Active/Inactive) ==========
        if user.role == 'teacher':
            # Get active academic year
            active_year = AcademicYear.objects.filter(is_active=True).first()
            
            if active_year:
                # Get teacher profile
                teacher_profile = TeacherProfile.objects.filter(user=user).first()
                
                if teacher_profile:
                    # Check if teacher has an enrollment for the active year
                    teacher_enrollment = Enrollment.objects.filter(
                        class_teacher=teacher_profile,
                        academic_year=active_year,
                        student__isnull=True
                    ).first()
                    
                    # If enrollment exists and status is Inactive, block login
                    if teacher_enrollment and teacher_enrollment.status == 'Inactive':
                        messages.error(
                            request, 
                            f" your account is INACTIVE for the current academic year "
                            f"({active_year.year_start}/{active_year.year_end}).\n\n"
                            f"contact school admin."
                        )
                        return redirect('login')
                    
                    # If no enrollment found for active year, also block login
                    if not teacher_enrollment:
                        messages.error(
                            request,
                            f" your account is not enrolled for the current academic year "
                            f"({active_year.year_start}/{active_year.year_end}).\n\n"
                            f"please contact admin"
                        )
                        return redirect('login')
                else:
                    messages.error(request, "your not registered yet contact admin for more information")
                    return redirect('login')
            else:
                messages.error(request, "no active academic year found")
                return redirect('login')
        
        # ========== CHECK STUDENT STATUS (Optional) ==========
        # if user.role == 'student':
        #     active_year = AcademicYear.objects.filter(is_active=True).first()
        #     if active_year:
        #         student_profile = StudentProfile.objects.filter(user=user).first()
        #         if student_profile:
        #             student_enrollment = Enrollment.objects.filter(
        #                 student=student_profile,
        #                 academic_year=active_year
        #             ).first()
        #             if student_enrollment and student_enrollment.status == 'Inactive':
        #                 messages.error(request, "Your account is inactive. Please contact the administrator.")
        #                 return redirect('login')

        # Update role if superuser
        if user.is_superuser and user.role != 'admin':
            user.role = 'admin'
            user.save()

        # Login user
        login(request, user)

        # Determine redirect URL based on role
        role_redirects = {
            'admin': 'admin_dashboard',
            'teacher': 'teacher_dashboard',
            'student': 'student_dashboard',
        }

        if user.role in role_redirects:
            messages.success(request, f"your most welcome, {user.get_full_name() or user.username}!")
            return redirect(role_redirects[user.role])

        # Invalid role
        logout(request)
        messages.error(request, "Invalid role!")
        return redirect('login')

    # GET request - render login page
    return render(request, 'attendance_app/login.html', {
        'admin_exists': admin_exists
    })
    
    
def format_phone_number(phone):
    if not phone:
        return None

    phone = re.sub(r"[^\d+]", "", phone)

    if phone.startswith("0") and len(phone) == 10:
        phone = "+255" + phone[1:]
    elif phone.startswith(("6", "7")) and len(phone) == 9:
        phone = "+255" + phone
    elif phone.startswith("255") and len(phone) == 12:
        phone = "+" + phone
    elif phone.startswith("+255") and len(phone) == 13:
        pass
    else:
        return None

    if phone.startswith(("+2556", "+2557")) and len(phone) == 13:
        return phone

    return None


@never_cache
def register_admin(request):
    if User.objects.filter(role='admin').exists():
        messages.warning(request, "Admin already exists! Please login.")
        return redirect('login')

    if request.method == "POST":
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip().lower()
        phone_number = request.POST.get('phone_number', '').strip()
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')

        if not all([first_name, last_name, email, phone_number, password1, password2]):
            messages.error(request, "All fields are required!")
            return redirect('register_admin')

        if password1 != password2:
            messages.error(request, "Passwords do not match!")
            return redirect('register_admin')

        if User.objects.filter(email=email).exists():
            messages.error(request, "Email already exists!")
            return redirect('register_admin')

        formatted_phone = format_phone_number(phone_number)
        
        if User.objects.filter(phone_number=formatted_phone).exists():
            messages.error(request, "Phone already used!")
            return redirect('register_admin')

        admin_user = User.objects.create(
            username=email,
            email=email,
            first_name=first_name,
            last_name=last_name,
            phone_number=formatted_phone,
            password=make_password(password1),
            role='admin',
            is_staff=True,
            is_superuser=True
        )

        TeacherProfile.objects.create(user=admin_user)

        messages.success(request, "Admin created successfully!")
        return redirect('login')

    return render(request, 'attendance_app/register_admin.html')


@never_cache  
def logout_view(request):
    logout(request)
    messages.success(request, "Logged out successfully.")
    return redirect('login')


# ================= ADMIN DASHBOARD VIEWS =================

@never_cache
@login_required
@user_passes_test(lambda u: u.role == 'admin')
def admin_dashboard(request):
    try:
        # Get active academic year
        active_year = AcademicYear.objects.filter(is_active=True).first()
        
        # ========== COUNT STUDENTS IN ACTIVE YEAR ONLY ==========
        if active_year:
            total_students = Enrollment.objects.filter(
                academic_year=active_year,
                status='Active',
                student__isnull=False
            ).values('student').distinct().count()
        else:
            total_students = 0
        
        # ========== COUNT TEACHERS IN ACTIVE YEAR ONLY ==========
        if active_year:
            total_teachers = Enrollment.objects.filter(
                academic_year=active_year,
                class_teacher__isnull=False
            ).values('class_teacher').distinct().count()
        else:
            total_teachers = 0
        
        # ========== COUNT CLASSROOMS IN ACTIVE YEAR ONLY ==========
        if active_year:
            classrooms_count = Classroom.objects.filter(year=active_year).count()
        else:
            classrooms_count = 0
        
        # ========== GET TEACHERS WITH THEIR ASSIGNMENTS ==========
        teachers_list = []
        if active_year:
            teacher_assignments = Enrollment.objects.filter(
                academic_year=active_year,
                class_teacher__isnull=False
            ).select_related('class_teacher__user', 'classroom', 'stream')
            
            for assignment in teacher_assignments:
                teacher = assignment.class_teacher
                teachers_list.append({
                    'id': teacher.id,
                    'name': teacher.user.get_full_name(),
                    'username': teacher.user.username,
                    'classroom': assignment.classroom.name if assignment.classroom else '—',
                    'stream': assignment.stream.name if assignment.stream else '—',
                })
        
        # Get recent activities
        recent_students = []
        if active_year:
            recent_enrollments = Enrollment.objects.filter(
                academic_year=active_year,
                status='Active',
                student__isnull=False
            ).select_related('student__user').order_by('-id')[:5]
            
            for enrollment in recent_enrollments:
                if enrollment.student and enrollment.student.user:
                    recent_students.append({
                        'name': enrollment.student.user.get_full_name(),
                        'admission_number': enrollment.student.admission_number,
                        'date': enrollment.enrollment_date if hasattr(enrollment, 'enrollment_date') else enrollment.academic_year.created_at,
                    })
        
        context = {
            'active_year': active_year,
            'total_students': total_students,
            'total_teachers': total_teachers,
            'classrooms_count': classrooms_count,
            'teachers': teachers_list,
            'recent_students': recent_students,
        }
        
        return render(request, 'attendance_app/admin_dashboard.html', context)
        
    except Exception as e:
        logger.error(f"Error in admin_dashboard: {e}")
        messages.error(request, f"An error occurred while loading dashboard: {str(e)}")
        return redirect('login')


# ================= TEACHER MANAGEMENT VIEWS =================

@never_cache
@login_required
@user_passes_test(lambda u: u.role == 'admin')
def manage_teacher(request):
    try:
        academic_years = AcademicYear.objects.all().order_by('-year_start')
        selected_year_id = request.GET.get('year')
        
        if selected_year_id:
            selected_academic_year = AcademicYear.objects.filter(id=selected_year_id).first()
        else:
            selected_academic_year = AcademicYear.objects.filter(is_active=True).first()
        
        teachers_list = []
        
        if selected_academic_year:
            enrollments = Enrollment.objects.filter(
                academic_year=selected_academic_year,
                student__isnull=True
            ).select_related(
                'class_teacher__user', 
                'classroom', 
                'stream', 
                'academic_year'
            )
            
            for enrollment in enrollments:
                teacher = enrollment.class_teacher
                if teacher:
                    teachers_list.append({
                        'teacher_id': teacher.id,
                        'username': teacher.user.username,
                        'full_name': teacher.user.get_full_name(),
                        'gender': teacher.user.gender if teacher.user.gender else '—',
                        'phone': teacher.user.phone_number or '—',
                        'email': teacher.user.email or '—',  # ADD THIS
                        'classroom': enrollment.classroom.name if enrollment.classroom else None,
                        'stream': enrollment.stream.name if enrollment.stream else None,
                        'academic_year': str(selected_academic_year),
                        'is_active': selected_academic_year.is_active,
                        'teacher_status': enrollment.status,
                    })
        
        paginator = Paginator(teachers_list, 25)
        page_number = request.GET.get('page')
        teachers = paginator.get_page(page_number)
        
        return render(request, 'attendance_app/manage_teacher.html', {
            'teachers': teachers,
            'academic_years': academic_years,
            'selected_academic_year': selected_academic_year,
        })
        
    except Exception as e:
        logger.error(f"Error in manage_teacher: {e}")
        messages.error(request, f"An error occurred while loading teachers: {str(e)}")
        return redirect('admin_dashboard')


@never_cache
@login_required
@user_passes_test(lambda u: u.role == 'admin')
def register_teacher(request):
    try:
        # Get active academic year
        active_year = AcademicYear.objects.filter(is_active=True).first()
        
        if not active_year:
            messages.error(request, "No active academic year found. Please activate an academic year first.")
            return redirect('academic_years')
        
        # Get ONLY classrooms for the active academic year
        classrooms = Classroom.objects.filter(year=active_year).order_by('name')

        if request.method == 'POST':
            email = request.POST.get('email', '').strip()
            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            gender = request.POST.get('gender')
            phone_number = request.POST.get('phone_number', '').strip()
            classroom_id = request.POST.get('classroom')
            stream_id = request.POST.get('stream')

            # ===== EMAIL VALIDATION =====
            if not email or '@' not in email:
                messages.error(request, "Please enter a valid email.")
                return redirect('register_teacher')

            # ===== CLASSROOM =====
            try:
                classroom = Classroom.objects.get(id=classroom_id, year=active_year)
            except Classroom.DoesNotExist:
                messages.error(request, "Classroom not found for the active academic year.")
                return redirect('register_teacher')

            # ===== STREAM =====
            stream = None
            if stream_id:
                try:
                    stream = Stream.objects.get(id=stream_id, classroom=classroom)
                    if Enrollment.objects.filter(stream=stream, academic_year=active_year, class_teacher__isnull=False).exists():
                        messages.warning(request, "This stream already has a teacher assigned for the active academic year.")
                        return redirect('register_teacher')
                except Stream.DoesNotExist:
                    messages.error(request, "Stream not found.")
                    return redirect('register_teacher')

            # ===== USER EXISTS =====
            if User.objects.filter(email=email).exists():
                messages.error(request, "An account with this email already exists.")
                return redirect('register_teacher')

            # ===== PHONE NORMALIZATION =====
            if phone_number:
                formatted_phone = format_phone_number(phone_number)
                if not formatted_phone:
                    messages.error(request, "Invalid phone number format! Use: 0712345678 or 712345678")
                    return redirect('register_teacher')
                phone_number = formatted_phone

            # ===== CREATE USER =====
            try:
                user = User.objects.create(
                    username=email,
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    password=make_password(DEFAULT_PASSWORD),
                    role='teacher',
                    gender=gender,
                    phone_number=phone_number
                )

                teacher = TeacherProfile.objects.create(user=user)

                Enrollment.objects.create(
                    classroom=classroom,
                    stream=stream,
                    class_teacher=teacher,
                    academic_year=active_year
                )

            except IntegrityError as e:
                messages.error(request, f"Database error: {str(e)}")
                return redirect('register_teacher')

            # ===== SEND SMS =====
            if phone_number:
                try:
                    africastalking.initialize(
                        username=settings.AFRICASTALKING_USERNAME,
                        api_key=settings.AFRICASTALKING_API_KEY
                    )
                    sms = africastalking.SMS
                    sms.send(
                        message=(
                            f"Habari {first_name}, account yako "
                            f"imesajiliwa kwa .\nEmail: {email}\nPassword: {DEFAULT_PASSWORD}"
                        ),
                        recipients=[phone_number],
                        sender_id='School_SMS'
                    )
                except Exception as e:
                    logger.error(f"SMS failed: {e}")

            messages.success(
                request,
                f"Teacher {first_name} {last_name} registered successfully for {active_year.year_start}/{active_year.year_end}!"
            )
            return redirect('register_teacher')

        return render(
            request,
            'attendance_app/register_teacher.html',
            {
                'classrooms': classrooms,
                'active_year': active_year,
            }
        )
        
    except Exception as e:
        logger.error(f"Error in register_teacher: {e}")
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect('manage_teacher')
    

def get_streams(request, classroom_id):
    streams = Stream.objects.filter(classroom_id=classroom_id)
    available_only = request.GET.get('available_only')
    if available_only == '1':
        active_year = AcademicYear.objects.filter(is_active=True).first()
        streams = streams.exclude(
            class_enrollments__class_teacher__isnull=False,
            class_enrollments__academic_year=active_year
        )

    data = {'streams': [{'id': s.id, 'name': s.name} for s in streams]}
    return JsonResponse(data)


@never_cache
@login_required
@user_passes_test(lambda u: u.role == 'admin')
def edit_teacher(request, id):
    try:
        # Optimize: include user in same query
        teacher = get_object_or_404(
            TeacherProfile.objects.select_related('user'),
            id=id
        )

        # Get active academic year
        active_year = AcademicYear.objects.filter(is_active=True).only(
            'id', 'year_start', 'year_end'
        ).first()

        if not active_year:
            messages.error(request, "No active academic year found. Please activate an academic year first.")
            return redirect('manage_teacher')

        # Optimized classrooms query (limit fields)
        classrooms = Classroom.objects.filter(
            year=active_year
        ).only('id', 'name').order_by('name')

        # Get current enrollment (optimized)
        enrollment = teacher.class_enrollments.filter(
            academic_year=active_year
        ).select_related('classroom', 'stream').first()

        current_classroom = enrollment.classroom if enrollment else None
        current_stream = enrollment.stream if enrollment else None

        if request.method == 'POST':
            email = request.POST.get('email', '').strip()
            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            password = request.POST.get('password', '').strip()
            phone_number = request.POST.get('phone_number', '').strip()
            gender = request.POST.get('gender')
            classroom_id = request.POST.get('classroom')
            stream_id = request.POST.get('stream')
            
            if phone_number:
               formatted_phone = format_phone_number(phone_number)
               if not formatted_phone:
                    messages.error(request, "Invalid phone number format!")
                    return redirect('edit_teacher', id=teacher.id)
               phone_number = formatted_phone

            try:
                with transaction.atomic():
                    # Update user (already fetched via select_related)
                    user = teacher.user
                    user.email = email
                    user.username = email
                    user.first_name = first_name
                    user.last_name = last_name
                    user.phone_number = phone_number
                    user.gender = gender
                    if password:
                       user.set_password(password)
                    user.full_clean()
                    user.save()

                    # Optimized classroom fetch (no exception)
                    classroom = None
                    if classroom_id:
                        classroom = Classroom.objects.filter(
                            id=classroom_id,
                            year=active_year
                        ).only('id').first()

                        if not classroom:
                            messages.error(request, "Selected classroom is not available for the active academic year.")
                            return redirect('edit_teacher', id=teacher.id)

                    # Optimized stream fetch
                    stream = None
                    if stream_id and classroom:
                        stream = Stream.objects.filter(
                            id=stream_id,
                            classroom=classroom
                        ).only('id').first()

                        if not stream:
                            messages.warning(request, "Selected stream not found for this classroom.")

                    # Update or create enrollment
                    if enrollment:
                        enrollment.classroom = classroom
                        enrollment.stream = stream
                        enrollment.save(update_fields=['classroom', 'stream'])
                    else:
                        Enrollment.objects.create(
                            classroom=classroom,
                            stream=stream,
                            class_teacher=teacher,
                            academic_year=active_year,
                            student=None
                        )

                messages.success(
                    request,
                    f"Teacher {user.get_full_name()} updated successfully for "
                    f"{active_year.year_start}/{active_year.year_end}!"
                )
                return redirect('manage_teacher')

            except ValidationError as e:
                messages.error(request, f"Validation Error: {e}")
                return redirect('edit_teacher', id=teacher.id)
            except Exception as e:
                messages.error(request, f"An unexpected error occurred: {str(e)}")
                return redirect('edit_teacher', id=teacher.id)

        # ---------------------------
        # STREAMS (OPTIMIZED QUERY)
        # ---------------------------
        streams = Stream.objects.none()

        if current_classroom:
            # Subquery instead of heavy JOIN
            assigned_stream_ids = Enrollment.objects.filter(
                academic_year=active_year,
                class_teacher__isnull=False
            ).values_list('stream_id', flat=True)

            streams = Stream.objects.filter(
                classroom=current_classroom
            ).exclude(
                id__in=assigned_stream_ids
            )

            # Ensure current stream is included
            if current_stream:
                streams = (streams | Stream.objects.filter(id=current_stream.id)).distinct()

        return render(request, 'attendance_app/edit_teacher.html', {
            'teacher': teacher,
            'classrooms': classrooms,
            'streams': streams,
            'current_classroom': current_classroom,
            'current_stream': current_stream,
            'active_year': active_year,
            'enrollment': enrollment,
        })
        
    except Exception as e:
        logger.error(f"Error in edit_teacher: {e}")
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect('manage_teacher')



@never_cache
@login_required
@user_passes_test(lambda u: u.role == 'admin')
def delete_teacher(request, teacher_id):
    """Delete teacher - Supports soft delete (toggle active/inactive), remove from year, or permanent delete"""
    
    teacher = get_object_or_404(
        TeacherProfile.objects.select_related('user'),
        id=teacher_id
    )
    
    teacher_name = teacher.user.get_full_name() or teacher.user.username
    
    # Get parameters
    action = request.GET.get('action', 'soft')
    selected_year_id = request.GET.get('year')
    toggle_action = request.GET.get('toggle', '')  # 'activate' or 'deactivate'
    
    # ================= TOGGLE ACTIVE/INACTIVE =================
    if toggle_action in ['activate', 'deactivate']:
        if not selected_year_id:
            messages.error(request, "Please select an academic year to change teacher status.")
            return redirect('manage_teacher')
        
        academic_year = get_object_or_404(AcademicYear, id=selected_year_id)
        
        # Find enrollment for the selected year
        enrollment = Enrollment.objects.filter(
            class_teacher=teacher,
            academic_year=academic_year,
            student__isnull=True
        ).first()
        
        if enrollment:
            if toggle_action == 'activate':
                enrollment.status = 'Active'
                enrollment.save()
                messages.success(
                    request,
                    f"Teacher {teacher_name} has been REACTIVATED in {academic_year.year_start}/{academic_year.year_end}!"
                )
            else:  # deactivate
                enrollment.status = 'Inactive'
                enrollment.save()
                messages.success(
                    request,
                    f"Teacher {teacher_name} has been DEACTIVATED in {academic_year.year_start}/{academic_year.year_end}."
                )
        else:
            messages.warning(
                request,
                f"No assignment found for {teacher_name} in {academic_year.year_start}/{academic_year.year_end}."
            )
        
        return redirect('manage_teacher')
    
    # ================= PERMANENT DELETE =================
    if action == 'permanent' or request.GET.get('permanent') == 'true':
        try:
            with transaction.atomic():
                enrollment_count = Enrollment.objects.filter(class_teacher=teacher).count()
                Enrollment.objects.filter(class_teacher=teacher).delete()
                user_to_delete = teacher.user
                user_to_delete.delete()
                
                messages.success(
                    request,
                    f"Teacher {teacher_name} PERMANENTLY deleted!\n"
                    f"Removed from {enrollment_count} class assignment(s)."
                )
                
                if request.user == user_to_delete:
                    logout(request)
                    return redirect('login')
                    
        except Exception as e:
            logger.error(f"Error permanently deleting teacher {teacher_id}: {e}")
            messages.error(request, f" Failed to permanently delete teacher: {str(e)}")
            
        return redirect('manage_teacher')
    
    # ================= SOFT DELETE (Remove from specific year only) =================
    else:
        if not selected_year_id:
            messages.error(request, "Please select an academic year to remove teacher from.")
            return redirect('manage_teacher')
        
        academic_year = get_object_or_404(AcademicYear, id=selected_year_id)
        
        deleted_count, _ = Enrollment.objects.filter(
            class_teacher=teacher,
            academic_year=academic_year,
            student__isnull=True
        ).delete()
        
        if deleted_count > 0:
            messages.success(
                request,
                f"Teacher {teacher_name} removed from "
                f"{academic_year.year_start}/{academic_year.year_end} successfully."
            )
        else:
            messages.warning(
                request,
                f"Teacher {teacher_name} has no assignment in {academic_year.year_start}/{academic_year.year_end}."
            )
        
        return redirect('manage_teacher')
    
    
# ================= STUDENT MANAGEMENT VIEWS =================

@never_cache
@login_required
@user_passes_test(lambda u: u.role in ['teacher', 'admin'])
def register_student_admin(request):
    try:

            # ONLY ACTIVE YEAR
            active_year = AcademicYear.objects.filter(is_active=True).first()

            if not active_year:
                messages.error(request, "No active academic year found!")
                return redirect('register_student_admin')

            # ONLY CLASSROOMS OF ACTIVE YEAR
            classrooms = Classroom.objects.filter(year=active_year).order_by('name')

            if request.method == 'POST':

                first_name = request.POST.get('first_name', '').strip()
                last_name = request.POST.get('last_name', '').strip()
                gender = request.POST.get('gender')
                password = request.POST.get('password') or "Student@123"
                admission_number = request.POST.get('admission_number', '').strip()
                classroom_id = request.POST.get('classroom')
                stream_id = request.POST.get('stream')
                parent_full_name = request.POST.get('parent_full_name', '').strip()
                parent_phone = request.POST.get('parent_phone', '').strip()

                # VALIDATIONS
                if not all([first_name, last_name, admission_number, classroom_id, stream_id, parent_full_name, parent_phone]):
                    messages.error(request, "All fields are required!")
                    return redirect('register_student_admin')

                # Format parent phone
                parent_phone = format_phone_number(parent_phone)
                if not parent_phone:
                    messages.error(request, "Invalid parent phone number format! Use: 0712345678 or 712345678")
                    return redirect('register_student_admin')

                # Check if admission number exists
                if User.objects.filter(username=admission_number).exists():
                    messages.error(request, "Admission number already exists.")
                    return redirect('register_student_admin')

                try:
                    with transaction.atomic():
                        # Get classroom and stream
                        classroom = Classroom.objects.get(id=classroom_id, year=active_year)
                        stream = Stream.objects.get(id=stream_id, classroom=classroom)

                        # CREATE STUDENT USER
                        student_user = User.objects.create(
                            username=admission_number,
                            first_name=first_name,
                            last_name=last_name,
                            gender=gender,
                            role='student',
                            password=make_password(password)
                        )

                        # CREATE STUDENT PROFILE
                        student_profile = StudentProfile.objects.create(
                            user=student_user,
                            admission_number=admission_number
                        )

                        # CREATE ENROLLMENT
                        Enrollment.objects.create(
                            student=student_profile,
                            classroom=classroom,
                            stream=stream,
                            academic_year=active_year,
                            status='Active'
                        )

                        # CREATE OR GET PARENT
                        parent_names = parent_full_name.split(' ', 1)
                        parent_first = parent_names[0]
                        parent_last = parent_names[1] if len(parent_names) > 1 else ''

                        # Get or create parent user
                        parent_user, created = User.objects.get_or_create(
                            username=parent_phone,
                            defaults={
                                'first_name': parent_first,
                                'last_name': parent_last,
                                'phone_number': parent_phone,
                                'role': 'parent',
                                'password': make_password(parent_phone)
                            }
                        )

                        # If user already exists, update their info
                        if not created:
                            parent_user.first_name = parent_first
                            parent_user.last_name = parent_last
                            parent_user.phone_number = parent_phone
                            parent_user.save()

                        # Create parent profile
                        ParentProfile.objects.get_or_create(
                            user=parent_user,
                            student=student_profile
                        )

                        messages.success(
                            request,
                            f"Student {first_name} {last_name} registered successfully for {active_year.year_start}/{active_year.year_end}!\n"
                            f"Login: {admission_number} | Password: {password}"
                        )
                        return redirect('register_student_admin')

                except Classroom.DoesNotExist:
                    messages.error(request, "Classroom not found for the active academic year.")
                    return redirect('register_student_admin')
                except Stream.DoesNotExist:
                    messages.error(request, "Stream not found.")
                    return redirect('register_student_admin')
                except IntegrityError as e:
                    messages.error(request, f"Database error: {str(e)}")
                    return redirect('register_student_admin')
                except Exception as e:
                    logger.error(f"Error registering student: {e}")
                    messages.error(request, f"Error: {str(e)}")
                    return redirect('register_student_admin')

            return render(request, 'attendance_app/register_student_admin.html', {
                'classrooms': classrooms,
                'active_year': active_year,
            })
            
    except Exception as e:
        logger.error(f"Error in register_student_admin: {e}")
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect('admin_dashboard')
    
    
    
@never_cache
@login_required
@user_passes_test(lambda u: u.role in ['teacher', 'admin'])
def manage_student(request):
    try:
        # Get all academic years for dropdown
        academic_years = AcademicYear.objects.all().order_by('-year_start')
        
        # Get selected academic year from request
        selected_year_id = request.GET.get('year')
        selected_academic_year = None
        
        if selected_year_id:
            try:
                selected_academic_year = AcademicYear.objects.get(id=selected_year_id)
            except AcademicYear.DoesNotExist:
                selected_academic_year = AcademicYear.objects.filter(is_active=True).first()
        else:
            selected_academic_year = AcademicYear.objects.filter(is_active=True).first()
        
        # Get selected classroom
        selected_classroom_id = request.GET.get('classroom')
        selected_classroom = None
        students_data = []
        available_streams = []
        search_query = request.GET.get('search', '')
        selected_stream_id = request.GET.get('stream', '')
        students = []
        
        if selected_academic_year:
            # Get classrooms for selected academic year only
            classrooms = Classroom.objects.filter(year=selected_academic_year).order_by('name')
            
            # Add student count to each classroom (include ALL statuses for counting)
    # Count ONLY students for this academic year
            for classroom in classrooms:
                classroom.student_count = Enrollment.objects.filter(
                    classroom=classroom,
                    academic_year=selected_academic_year,
                    student__isnull=False
                ).count()
            
            # If a classroom is selected, get its students
            if selected_classroom_id:
                try:
                    selected_classroom = Classroom.objects.get(id=selected_classroom_id, year=selected_academic_year)
                    
                    # Get available streams for this classroom
                    available_streams = Stream.objects.filter(classroom=selected_classroom)
                    
                    # IMPORTANT FIX: Get ALL enrollments for this classroom and year (NOT just Active)
                    enrollments = Enrollment.objects.filter(
                        classroom=selected_classroom,
                        academic_year=selected_academic_year,
                        student__isnull=False
                        # REMOVED: status='Active' - This was hiding past data!
                    ).select_related('student__user', 'stream')
                    
                    # Apply stream filter if selected
                    if selected_stream_id:
                        enrollments = enrollments.filter(stream_id=selected_stream_id)
                    
                    # Build student data
                    for enrollment in enrollments:
                        student = enrollment.student
                        
                        # Skip if student or user is None
                        if not student or not student.user:
                            continue
                        
                        full_name = f"{student.user.first_name} {student.user.last_name}".strip()
                        if not full_name or not student.user.first_name:
                            full_name = student.user.username or "Unknown Student"
                        
                        # Apply search filter
                        if search_query:
                            search_lower = search_query.lower()
                            if search_lower not in full_name.lower() and search_lower not in student.admission_number.lower():
                                continue
                        
                        students_data.append({
                            'id': student.id,
                            'admission_number': student.admission_number,
                            'full_name': full_name,
                            'gender': student.user.gender or '—',
                            'stream': enrollment.stream.name if enrollment.stream else '—',
                            'status': enrollment.status,  # This will show 'Promoted', 'Graduated', or 'Active'
                        })
                    
                    # Pagination
                    if students_data:
                        paginator = Paginator(students_data, 25)
                        page_number = request.GET.get('page')
                        students = paginator.get_page(page_number)
                    
                except Classroom.DoesNotExist:
                    pass
        else:
            classrooms = []

        context = {
            'academic_years': academic_years,
            'selected_academic_year': selected_academic_year,
            'classrooms': classrooms,
            'selected_classroom': selected_classroom,
            'students': students,
            'available_streams': available_streams,
            'search_query': search_query,
            'selected_stream_id': selected_stream_id,
        }
        
        return render(request, 'attendance_app/manage_student.html', context)
    except Exception as e:
        logger.error(f"Error in manage_student: {e}")
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect('admin_dashboard')
    


@never_cache
@login_required
@user_passes_test(lambda u: u.role in ['admin'])
def classroom_students(request, classroom_id):
    """View students in a specific classroom"""
    
    try:
        # Get the classroom
        classroom = get_object_or_404(Classroom, id=classroom_id)
        
        # Get academic year from URL or use classroom's year
        selected_year_id = request.GET.get('year')
        if selected_year_id:
            academic_year = get_object_or_404(AcademicYear, id=selected_year_id)
        else:
            academic_year = classroom.year
        
        # Get streams for this classroom
        streams = Stream.objects.filter(classroom=classroom)
        
        # Get filter parameters
        search_query = request.GET.get('search', '')
        selected_stream_id = request.GET.get('stream', '')
        
        # Get ALL enrollments for this classroom and year
        enrollments = Enrollment.objects.filter(
            classroom=classroom,
            academic_year=academic_year,
            student__isnull=False
        ).select_related('student__user', 'stream')
        
        # Apply stream filter
        if selected_stream_id:
            enrollments = enrollments.filter(stream_id=selected_stream_id)
        
        # Build student data with parent info
        students_data = []
        for enrollment in enrollments:
            student = enrollment.student
            if not student or not student.user:
                continue
            
            full_name = f"{student.user.first_name} {student.user.last_name}".strip()
            if not full_name or not student.user.first_name:
                full_name = student.user.username or "Unknown Student"
            
            # Get parent info
            parent = student.parents.first()
            parent_name = f"{parent.user.first_name} {parent.user.last_name}".strip() if parent and parent.user else '—'
            parent_phone = parent.user.phone_number if parent and parent.user and parent.user.phone_number else '—'
            
            # Apply search filter
            if search_query:
                search_lower = search_query.lower()
                if search_lower not in full_name.lower() and search_lower not in student.admission_number.lower():
                    continue
            
            students_data.append({
                'id': student.id,
                'admission_number': student.admission_number,
                'full_name': full_name,
                'gender': student.user.gender or '—',
                'stream': enrollment.stream.name if enrollment.stream else '—',
                'status': enrollment.status,
                'email': student.user.email or '—',
                'phone': student.user.phone_number or '—',
                'parent_name': parent_name,
                'parent_phone': parent_phone,
            })
        
        # Pagination
        paginator = Paginator(students_data, 25)
        page_number = request.GET.get('page')
        students = paginator.get_page(page_number)
        
        context = {
            'classroom': classroom,
            'academic_year': academic_year,
            'streams': streams,
            'students': students,
            'search_query': search_query,
            'selected_stream_id': selected_stream_id,
            'selected_academic_year': academic_year,
        }
        
        return render(request, 'attendance_app/classroom_students.html', context)

    except Exception as e:
        logger.error(f"Error in classroom_students: {e}")
        messages.error(request, f"An error occurred while loading students: {str(e)}")
        return redirect('manage_student')
    

@never_cache
@login_required
@user_passes_test(lambda u: u.role in ['teacher'])
def register_student_teacher(request):
    try:
        teacher = get_object_or_404(TeacherProfile, user=request.user)
        active_year = AcademicYear.objects.filter(is_active=True).first()
        
        if not active_year:
            messages.error(request, "No active academic year found. Please contact admin.")
            return redirect('teacher_dashboard')
        
        teacher_enrollment = Enrollment.objects.filter(
            class_teacher=teacher,
            academic_year=active_year
        ).select_related('classroom', 'stream').first()
        
        if not teacher_enrollment:
            messages.error(request, "You are not assigned to any classroom for the current academic year.")
            return redirect('teacher_dashboard')
        
        classroom = teacher_enrollment.classroom
        stream = teacher_enrollment.stream
        
        if not classroom:
            messages.error(request, "You are not properly assigned to a classroom.")
            return redirect('teacher_dashboard')
        
        academic_years = AcademicYear.objects.all()

        students = StudentProfile.objects.filter(
            enrollments__classroom=classroom,
            enrollments__academic_year=active_year
        ).distinct()
        
        if stream:
            students = students.filter(enrollments__stream=stream)

        if request.method == 'POST':
            if 'import_excel' in request.POST and request.FILES.get('excel_file'):
                try:
                    # Read Excel file
                    excel_file = request.FILES['excel_file']
                    df = pd.read_excel(excel_file)
                    
                    # Required columns validation
                    required_columns = ['first_name', 'last_name', 'gender', 'admission_number']
                    missing_columns = [col for col in required_columns if col not in df.columns]
                    
                    if missing_columns:
                        messages.error(request, f"Missing required columns: {', '.join(missing_columns)}")
                        return redirect('register_student_teacher')
                    
                    success_count = 0
                    error_count = 0
                    error_messages = []
                    
                    # Use transaction for data integrity
                    with transaction.atomic():
                        for index, row in df.iterrows():
                            try:
                                admission_number = str(row['admission_number']).strip()
                                
                                # Check if user exists
                                if User.objects.filter(username=admission_number).exists():
                                    error_count += 1
                                    error_messages.append(f"Row {index+2}: Admission number {admission_number} already exists")
                                    continue
                                
                                # Get gender and validate
                                gender = row.get('gender', '').strip()
                                if gender not in ['Male', 'Female']:
                                    error_count += 1
                                    error_messages.append(f"Row {index+2}: Invalid gender '{gender}'. Must be 'Male' or 'Female'")
                                    continue
                                
                                # Create student user
                                student_user = User.objects.create(
                                    username=admission_number,
                                    first_name=str(row['first_name']).strip(),
                                    last_name=str(row['last_name']).strip(),
                                    gender=gender,
                                    role='student',
                                    password=make_password("Student@123")
                                )
                                
                                # Create student profile
                                student_profile = StudentProfile.objects.create(
                                    user=student_user,
                                    admission_number=admission_number
                                )
                                
                                # Create enrollment
                                Enrollment.objects.create(
                                    student=student_profile,
                                    classroom=classroom,
                                    stream=stream,
                                    academic_year=active_year,
                                    status='Active'
                                )
                                
                                # Handle parent if provided
                                parent_full_name = str(row.get('parent_full_name', '')).strip()
                                parent_phone = str(row.get('parent_phone', '')).strip()
                                
                                if parent_full_name and parent_phone:
                                    # TUMIA FORMAT_PHONE_NUMBER ILIYOPO KWENYE HELPER FUNCTIONS
                                    parent_phone_clean = format_phone_number(parent_phone)
                                    
                                    if parent_phone_clean:
                                        parent_names = parent_full_name.split(' ', 1)
                                        parent_first = parent_names[0]
                                        parent_last = parent_names[1] if len(parent_names) > 1 else ''
                                        
                                        parent_user, created = User.objects.get_or_create(
                                            username=parent_phone_clean,
                                            defaults={
                                                'first_name': parent_first,
                                                'last_name': parent_last,
                                                'phone_number': parent_phone_clean,
                                                'role': 'parent',
                                                'password': make_password(parent_phone_clean)
                                            }
                                        )
                                        
                                        ParentProfile.objects.get_or_create(
                                            user=parent_user,
                                            student=student_profile
                                        )
                                    else:
                                        error_messages.append(f"Row {index+2}: Invalid parent phone format for {parent_full_name}")
                                
                                success_count += 1
                                
                            except Exception as e:
                                error_count += 1
                                error_messages.append(f"Row {index+2}: {str(e)}")
                    
                    # Show results
                    if success_count > 0:
                        messages.success(request, f"Successfully imported {success_count} student(s)!")
                    
                    if error_messages:
                        # Limit to first 5 errors to avoid overwhelming
                        for msg in error_messages[:5]:
                            messages.warning(request, msg)
                        if len(error_messages) > 5:
                            messages.warning(request, f"... and {len(error_messages) - 5} more errors")
                    
                    return redirect('register_student_teacher')
                    
                except Exception as e:
                    messages.error(request, f"Error reading Excel file: {str(e)}")
                    return redirect('register_student_teacher')

            # Manual registration code
            admission_number = request.POST.get('admission_number')
            
            if not admission_number:
                messages.error(request, "Admission number is required.")
                return redirect('register_student_teacher')

            if User.objects.filter(username=admission_number).exists():
                messages.error(request, "Admission number already exists.")
                return redirect('register_student_teacher')

            student_user = User.objects.create(
                username=admission_number,
                first_name=request.POST.get('first_name'),
                last_name=request.POST.get('last_name'),
                gender=request.POST.get('gender'),
                role='student',
                password=make_password(request.POST.get('password') or "Student@123")
            )

            student_profile = StudentProfile.objects.create(
                user=student_user,
                admission_number=admission_number
            )

            Enrollment.objects.create(
                student=student_profile,
                classroom=classroom,
                stream=stream,
                academic_year=active_year,
                status='Active'
            )

            parent_full_name = request.POST.get('parent_full_name', '').strip()
            parent_phone = request.POST.get('parent_phone', '').strip()

            if parent_full_name and parent_phone:
                # phone number format
                parent_phone_clean = format_phone_number(parent_phone)
                
                if parent_phone_clean:
                    parent_names = parent_full_name.split(' ', 1)
                    parent_first = parent_names[0]
                    parent_last = parent_names[1] if len(parent_names) > 1 else ''

                    parent_user, created = User.objects.get_or_create(
                        username=parent_phone_clean,
                        defaults={
                            'first_name': parent_first,
                            'last_name': parent_last,
                            'phone_number': parent_phone_clean,
                            'role': 'parent',
                            'password': make_password(parent_phone_clean)
                        }
                    )

                    ParentProfile.objects.get_or_create(
                        user=parent_user,
                        student=student_profile
                    )
                else:
                    messages.warning(request, 'Invalid parent phone number format. Use: 0712345678 or 712345678. Student saved, but parent not linked.')

            messages.success(request, f"Student {student_user.get_full_name()} registered successfully.")
            return redirect('register_student_teacher')

        return render(request, 'attendance_app/register_student_teacher.html', {
            'students': students,
            'active_year': active_year,
            'academic_years': academic_years,
            'teacher_classroom': classroom,
            'teacher_stream': stream,
        })
    except Exception as e:
        logger.error(f"Error in register_student_teacher: {e}")
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect('teacher_dashboard')


@login_required
@user_passes_test(lambda u: u.role in ['teacher', 'admin'])
def edit_student_page(request, student_id):

    student = get_object_or_404(StudentProfile, id=student_id)
    parent = student.parents.first()
    active_year = AcademicYear.objects.filter(is_active=True).first()

    if not active_year:
        messages.error(request, "No active academic year found!")
        return redirect('manage_student')

    # GET CURRENT ENROLLMENT FOR ACTIVE YEAR
    enrollment = student.enrollments.filter(
        academic_year=active_year
    ).select_related('classroom', 'stream').first()

    # FETCH ONLY ACTIVE YEAR CLASSROOMS
    classrooms = Classroom.objects.filter(year=active_year).order_by('name')

    # GET STREAMS FOR CURRENT CLASSROOM
    streams = Stream.objects.none()
    if enrollment and enrollment.classroom:
        streams = Stream.objects.filter(classroom=enrollment.classroom).order_by('name')

    if request.method == 'POST':

        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        gender = request.POST.get('gender')
        admission_number = request.POST.get('admission_number', '').strip()
        classroom_id = request.POST.get('classroom')
        stream_id = request.POST.get('stream')
        parent_full_name = request.POST.get('parent_full_name', '').strip()
        parent_phone = request.POST.get('parent_phone', '').strip()

        # VALIDATIONS
        if not all([first_name, last_name, admission_number, parent_full_name, parent_phone]):
            messages.error(request, "Please fill all required fields!")
            return redirect('edit_student_page', student_id=student.id)

        # Format parent phone
        parent_phone = format_phone_number(parent_phone)
        if not parent_phone:
            messages.error(request, "Invalid parent phone number format!")
            return redirect('edit_student_page', student_id=student.id)

        try:
            with transaction.atomic():
                # UPDATE STUDENT USER
                student.user.first_name = first_name
                student.user.last_name = last_name
                student.user.gender = gender
                student.user.save()

                # UPDATE STUDENT PROFILE
                student.admission_number = admission_number
                student.save()

                # UPDATE ENROLLMENT
                if classroom_id:
                    new_classroom = Classroom.objects.filter(id=classroom_id, year=active_year).first()
                    if new_classroom:
                        if enrollment:
                            enrollment.classroom = new_classroom
                        else:
                            enrollment = Enrollment(
                                student=student,
                                classroom=new_classroom,
                                academic_year=active_year,
                                status='Active'
                            )
                
                if stream_id and enrollment:
                    new_stream = Stream.objects.filter(id=stream_id).first()
                    if new_stream:
                        enrollment.stream = new_stream
                
                if enrollment:
                    enrollment.save()

                # UPDATE OR CREATE PARENT
                parent_names = parent_full_name.split(' ', 1)
                parent_first = parent_names[0]
                parent_last = parent_names[1] if len(parent_names) > 1 else ''

                # Get or create parent user
                parent_user, created = User.objects.get_or_create(
                    username=parent_phone,
                    defaults={
                        'first_name': parent_first,
                        'last_name': parent_last,
                        'phone_number': parent_phone,
                        'role': 'parent',
                        'password': make_password(parent_phone)
                    }
                )

                # Update parent user info if not created or if info changed
                if not created or parent_user.first_name != parent_first:
                    parent_user.first_name = parent_first
                    parent_user.last_name = parent_last
                    parent_user.phone_number = parent_phone
                    parent_user.save()

                # Update or create parent profile
                if parent:
                    if parent.user != parent_user:
                        parent.user = parent_user
                        parent.save()
                else:
                    ParentProfile.objects.create(
                        user=parent_user,
                        student=student
                    )

                messages.success(request, " Student updated successfully!")
                return redirect('manage_student')

        except Classroom.DoesNotExist:
            messages.error(request, "Selected classroom not found.")
            return redirect('edit_student_page', student_id=student.id)
        except Stream.DoesNotExist:
            messages.error(request, "Selected stream not found.")
            return redirect('edit_student_page', student_id=student.id)
        except Exception as e:
            logger.error(f"Error editing student {student_id}: {e}")
            messages.error(request, f"Error: {str(e)}")
            return redirect('edit_student_page', student_id=student.id)

    context = {
        'student': student,
        'parent': parent,
        'classrooms': classrooms,
        'streams': streams,
        'enrollment': enrollment,
        'active_year': active_year,
    }
    return render(request, 'attendance_app/edit_student_page.html', context)


@login_required
@user_passes_test(lambda u: u.role in ['teacher', 'admin'])
def edit_student_teacher(request, student_id):
    try:
        student = get_object_or_404(StudentProfile, id=student_id)
        streams = Stream.objects.all()
        classrooms = Classroom.objects.all()
        parent = student.parents.first() if student.parents.exists() else None
        active_year = AcademicYear.objects.filter(is_active=True).first()

        if request.method == 'POST':
            student.user.first_name = request.POST.get('first_name', student.user.first_name)
            student.user.last_name = request.POST.get('last_name', student.user.last_name)
            student.user.gender = request.POST.get('gender', student.user.gender)
            student.admission_number = request.POST.get('admission_number', student.admission_number)
            student.user.save()
            student.save()

            # Update enrollment for class & stream (only admin)
            if request.user.role == 'admin':
                enrollment = student.enrollments.filter(academic_year=active_year).first()
                if enrollment:
                    classroom_id = request.POST.get('classroom')
                    if classroom_id:
                        enrollment.classroom = Classroom.objects.filter(id=classroom_id).first()
                    stream_id = request.POST.get('stream')
                    if stream_id:
                        enrollment.stream = Stream.objects.filter(id=stream_id).first()
                    enrollment.save()

            if parent:
                full_name = request.POST.get('parent_full_name', '').strip()
                if full_name:
                    names = full_name.split(' ', 1)
                    parent.user.first_name = names[0]
                    parent.user.last_name = names[1] if len(names) > 1 else ''

                raw_phone = request.POST.get('parent_phone')
                if raw_phone:
                    try:
                        parent.user.phone_number = format_phone_number(raw_phone)
                    except Exception:
                        pass

                parent.user.save()

            messages.success(request, "Student updated successfully.")
            return redirect('my_students')

        return render(request, 'attendance_app/edit_student_teacher.html', {
            'student': student,
            'streams': streams,
            'classrooms': classrooms,
            'parent': parent,
            'is_admin': request.user.role == 'admin',
        })

    except Exception as e:
        logger.error(f"Error in edit_student_teacher ({student_id}): {e}")
        messages.error(request, "Unable to update student. Please try again.")
        return redirect('my_students')

import logging
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.cache import never_cache
from django.db import transaction
from .models import StudentProfile, AcademicYear, Enrollment

logger = logging.getLogger(__name__)

@never_cache
@login_required
@user_passes_test(lambda u: u.role in ['teacher', 'admin'])
def delete_student(request, student_id):
    """Delete student - Supports soft delete (toggle active/inactive), remove from year, or permanent delete"""
    
    student = get_object_or_404(StudentProfile, id=student_id)
    student_name = student.user.get_full_name() or student.user.username
    
    # Get parameters
    action = request.GET.get('action', 'soft')
    selected_year_id = request.GET.get('year')
    toggle_action = request.GET.get('toggle', '')  # 'activate' or 'deactivate'
    
    # Log the request for debugging
    logger.info(f"delete_student called: student_id={student_id}, action={action}, year={selected_year_id}, toggle={toggle_action}")
    
    # ================= TOGGLE ACTIVE/INACTIVE =================
    if toggle_action in ['activate', 'deactivate']:
        if not selected_year_id:
            messages.error(request, "Please select an academic year to change student status.")
            if request.user.role == "teacher":
                return redirect("my_students")
            return redirect("manage_student")
        
        academic_year = get_object_or_404(AcademicYear, id=selected_year_id)
        
        # Find enrollment for the selected year
        enrollment = student.enrollments.filter(academic_year=academic_year).first()
        
        if enrollment:
            old_status = enrollment.status
            if toggle_action == 'activate':
                enrollment.status = 'Active'
                enrollment.save()
                messages.success(
                    request,
                    f" Student {student_name} has been REACTIVATED in {academic_year.year_start}/{academic_year.year_end}!"
                )
                logger.info(f"Student {student_name} reactivated: {old_status} -> Active")
            else:  # deactivate
                enrollment.status = 'Inactive'
                enrollment.save()
                messages.success(
                    request,
                    f" Student {student_name} has been DEACTIVATED in {academic_year.year_start}/{academic_year.year_end}."
                )
                logger.info(f"Student {student_name} deactivated: {old_status} -> Inactive")
        else:
            messages.warning(
                request,
                f"No enrollment found for {student_name} in {academic_year.year_start}/{academic_year.year_end}."
            )
        
        # Redirect back to where the request came from
        referer = request.META.get('HTTP_REFERER', 'manage_student')
        # Prevent redirect loop
        if referer and 'delete_student' in referer:
            referer = 'manage_student'
        return redirect(referer)
    
    # ================= PERMANENT DELETE (Remove all data) =================
    if action == 'permanent' or request.GET.get('permanent') == 'true':
        if request.user.role != 'admin':
            messages.error(request, "Only admin can permanently delete students.")
            return redirect("manage_student")
        
        try:
            with transaction.atomic():
                student_user = student.user
                
                # Count records before deletion
                sms_count = SMSLog.objects.filter(student=student).count()
                attendance_count = Attendance.objects.filter(student=student).count()
                parent_count = ParentProfile.objects.filter(student=student).count()
                enrollment_count = student.enrollments.count()
                
                # Delete all related data in correct order
                SMSLog.objects.filter(student=student).delete()
                ParentProfile.objects.filter(student=student).delete()
                Attendance.objects.filter(student=student).delete()
                student.enrollments.all().delete()
                
                # Delete student profile and user
                student.delete()
                student_user.delete()
                
                messages.success(
                    request, 
                    f" Student {student_name} PERMANENTLY deleted!\n"
                    f" Removed: {enrollment_count} enrollments, {attendance_count} attendance records, "
                    f"{parent_count} parent relationships, {sms_count} SMS logs."
                )
                
                logger.warning(f"PERMANENT DELETE: Student {student_name} (ID: {student_id}) deleted by admin {request.user.username}")
                
        except Exception as e:
            logger.error(f"Error in permanent delete: {e}")
            messages.error(request, f" Failed to permanently delete student: {str(e)}")
        
        return redirect("manage_student")
    
    # ================= SOFT DELETE (Remove from specific year only) =================
    else:
        if not selected_year_id:
            messages.error(request, "Please select an academic year to remove student from.")
            if request.user.role == "teacher":
                return redirect("my_students")
            return redirect("manage_student")
        
        academic_year = get_object_or_404(AcademicYear, id=selected_year_id)
        
        # Delete only enrollment for selected year
        deleted_count, _ = Enrollment.objects.filter(
            student=student,
            academic_year=academic_year
        ).delete()
        
        if deleted_count > 0:
            messages.success(
                request,
                f" Student {student_name} removed from {academic_year.year_start}/{academic_year.year_end} successfully."
            )
            logger.info(f"Student {student_name} removed from year {academic_year}")
        else:
            messages.warning(
                request,
                f" Student {student_name} has no enrollment in {academic_year.year_start}/{academic_year.year_end}."
            )
        
        return redirect("manage_student")
    
@never_cache
@login_required
def my_students(request):
    """OPTIMIZED - No N+1 queries"""
    try:
        teacher = get_object_or_404(TeacherProfile, user=request.user)
        active_year = AcademicYear.objects.filter(is_active=True).first()
        
        if not active_year:
            messages.error(request, "No active academic year found.")
            return redirect('teacher_dashboard')
        
        teacher_enrollment = Enrollment.objects.filter(
            class_teacher=teacher,
            academic_year=active_year
        ).select_related('classroom', 'stream').first()
        
        if not teacher_enrollment:
            messages.warning(request, "You are not assigned to any classroom.")
            return redirect('teacher_dashboard')
        
        classroom = teacher_enrollment.classroom
        stream = teacher_enrollment.stream
        search = request.GET.get("search", "")

        students_qs = StudentProfile.objects.filter(
            enrollments__classroom=classroom,
            enrollments__academic_year=active_year
        ).select_related('user').prefetch_related(
            'enrollments',
            'parents__user'
        ).distinct()

        if stream:
            students_qs = students_qs.filter(enrollments__stream=stream)
        
        if search:
            students_qs = students_qs.filter(
                Q(admission_number__icontains=search) |
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search)
            )

        students_qs = students_qs.order_by("user__first_name")
        total_count = students_qs.count()
        
        students_list = []
        for student in students_qs:
            enrollment = student.enrollments.filter(academic_year=active_year).first()
            parent = student.parents.first()
            
            # Get parent name
            parent_name = f"{parent.user.first_name} {parent.user.last_name}".strip() if parent and parent.user else '—'
            
            students_list.append({
                'id': student.id,
                'admission_number': student.admission_number,
                'full_name': student.user.get_full_name(),
                'classroom': enrollment.classroom.name if enrollment and enrollment.classroom else '—',
                'stream': enrollment.stream.name if enrollment and enrollment.stream else '—',
                'academic_year': str(enrollment.academic_year) if enrollment else '—',
                'gender': student.user.gender or '—',
                'phone': parent.user.phone_number if parent else '—',
                'parent_name': parent_name,  # ADD THIS LINE
            })

        paginator = Paginator(students_list, 25)
        page_number = request.GET.get("page")
        students = paginator.get_page(page_number)

        context = {
            "students": students,
            "classroom": classroom,
            "stream": stream,
            'teacher': teacher,
            "search": search,
            "active_year": active_year,
            "total_students": total_count,
        }
        return render(request, "attendance_app/my_students.html", context)

    except Exception as e:
        logger.error(f"Error in my_students: {e}")
        messages.error(request, "Unable to load students.")
        return redirect('teacher_dashboard')


# ================= TEACHER DASHBOARD & ATTENDANCE VIEWS =================

@never_cache
@login_required
def teacher_dashboard(request):
    try:
        """OPTIMIZED - Single query version"""
        teacher_profile = TeacherProfile.objects.select_related('user').get(user=request.user)
        active_year = AcademicYear.objects.filter(is_active=True).first()
        
        if not active_year:
            return render(request, 'attendance_app/teacher_dashboard.html', {'error': 'No active academic year'})
        
        # OPTIMIZED: Single query with all needed relations
        enrollment = Enrollment.objects.filter(
            class_teacher=teacher_profile,
            academic_year=active_year
        ).select_related('classroom', 'stream').first()

        classroom = enrollment.classroom if enrollment else None
        stream = enrollment.stream if enrollment else None

        # OPTIMIZED: Get all students in one query
        students = StudentProfile.objects.filter(
            enrollments__classroom=classroom,
            enrollments__academic_year=active_year
        ).select_related('user').distinct()
        
        if stream:
            students = students.filter(enrollments__stream=stream)
        
        total_students = students.count()
        student_ids = list(students.values_list('id', flat=True))
        
        # OPTIMIZED: Get attendance counts in one query using aggregation
        from django.db.models import Count, Q, Case, When, IntegerField, Sum
        
        attendance_counts = Attendance.objects.filter(
            student_id__in=student_ids
        ).aggregate(
            total=Count('id'),
            present=Sum(Case(When(status='present', then=1), default=0, output_field=IntegerField())),
            absent=Sum(Case(When(status='absent', then=1), default=0, output_field=IntegerField())),
            sick=Sum(Case(When(status='sick', then=1), default=0, output_field=IntegerField()))
        )
        
        # OPTIMIZED: Get parents in one query
        parents = ParentProfile.objects.filter(student_id__in=student_ids).select_related('user')
        
        # OPTIMIZED: Get recent SMS logs in one query
        sms_logs = SMSLog.objects.filter(student_id__in=student_ids).select_related(
            'student__user', 'parent__user'
        ).order_by('-timestamp')[:10]

        context = {
            'teacher': teacher_profile,
            'classroom': classroom,
            'stream': stream,
            'students': students,
            'parents': parents,
            'total_students': total_students,
            'total_attendance': attendance_counts['total'] or 0,
            'present_count': attendance_counts['present'] or 0,
            'absent_count': attendance_counts['absent'] or 0,
            'sick_count': attendance_counts['sick'] or 0,
            'sms_logs': sms_logs,
        }

        return render(request, 'attendance_app/teacher_dashboard.html', context)
    except Exception as e:
        logger.error(f"Error in teacher_dashboard: {e}")
        messages.error(request, "An error occurred loading dashboard")
        return redirect('login')


@never_cache
@login_required
@user_passes_test(lambda u: u.role == 'teacher')
def mark_attendance(request):
    try:
        teacher = get_object_or_404(TeacherProfile, user=request.user)
        active_year = AcademicYear.objects.filter(is_active=True).first()
        
        if not active_year:
            messages.error(request, "No active academic year found.")
            return redirect('teacher_dashboard')
        
        teacher_enrollment = Enrollment.objects.filter(
            class_teacher=teacher,
            academic_year=active_year
        ).select_related('classroom', 'stream').first()

        if not teacher_enrollment or not teacher_enrollment.classroom:
            messages.error(request, "You are not assigned to a classroom or stream.")
            return redirect('teacher_dashboard')
        
        classroom = teacher_enrollment.classroom
        stream = teacher_enrollment.stream

        students_list = get_students_by_teacher_scope(classroom, stream, academic_year=active_year).order_by('admission_number')
        paginator = Paginator(students_list, 25)
        page_number = request.GET.get('page')
        students = paginator.get_page(page_number)

        if request.method == "POST":
            today = timezone.localdate()
            
            # Track SMS results
            total_sms_sent = 0
            total_sms_failed = 0
            sms_errors = []
            low_balance_detected = False

            for student in students:
                status = request.POST.get(f'attendance_{student.id}', 'present')
                student_enrollment = Enrollment.objects.filter(student=student, academic_year=active_year).first()

                # Save attendance
                Attendance.objects.update_or_create(
                    student=student,
                    date=today,
                    enrollment=student_enrollment,
                    defaults={'status': status, 'marked_by': teacher}
                )

                # Send SMS only for absent students
                if status == 'absent':
                    # Check if SMS already sent today
                    start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
                    end = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
                    sms_exists = SMSLog.objects.filter(student=student, timestamp__range=(start, end)).exists()

                    if not sms_exists:
                        try:
                            success_count, failed_count, error_msgs = send_absent_sms(student, teacher=teacher)
                            total_sms_sent += success_count
                            total_sms_failed += failed_count
                            
                            # Check for low balance error
                            for msg in error_msgs:
                                if 'salio' in msg.lower() or 'balance' in msg.lower():
                                    low_balance_detected = True
                                sms_errors.append(f"{student.user.get_full_name()}: {msg}")
                                
                        except Exception as e:
                            total_sms_failed += 1
                            logger.error(f"Error sending absent SMS for {student.id}: {e}")
                            sms_errors.append(f"{student.user.get_full_name()}: System error")

            # Prepare feedback message
            feedback_parts = [f" Attendance marked successfully for {today.strftime('%d/%m/%Y')}"]
            
            if total_sms_sent > 0:
                feedback_parts.append(f" SMS sent: {total_sms_sent} parent(s)")
            
            if total_sms_failed > 0:
                feedback_parts.append(f" SMS failed: {total_sms_failed}")
            
            if low_balance_detected:
                feedback_parts.append(" ONYO: Salio la SMS limepungua au halipo!")
            
            # Show main success message
            messages.success(request, " | ".join(feedback_parts))
            
            # Show individual error details (max 3)
            for error in sms_errors[:3]:
                messages.warning(request, f"{error}")
            
            if len(sms_errors) > 3:
                messages.info(request, f"... na {len(sms_errors) - 3} errors nyingine")
            
            # Special alert for low balance
            if low_balance_detected:
                messages.error(request, " TAHADHARI: Salio la SMS halipo! Wasiliana na Admin kuongeza salio.")
            
            return redirect('view_attendance')

        context = {
            'students': students,
            'classroom': classroom,
            'stream': stream,
            'teacher': teacher,
        }
        return render(request, 'attendance_app/mark_attendance.html', context)
    except Exception as e:
        logger.error(f"Error in mark_attendance: {e}")
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect('teacher_dashboard')



@never_cache
@login_required
def view_attendance(request):
    try:
        """OPTIMIZED - Single query with aggregation"""
        teacher = get_object_or_404(TeacherProfile.objects.select_related('user'), user=request.user)
        active_year = AcademicYear.objects.filter(is_active=True).first()
        
        if not active_year:
            messages.error(request, "No active academic year found.")
            return redirect('teacher_dashboard')
        
        teacher_enrollment = Enrollment.objects.filter(
            class_teacher=teacher,
            academic_year=active_year
        ).select_related('classroom', 'stream').first()
        
        if not teacher_enrollment or not teacher_enrollment.classroom:
            messages.error(request, "You are not assigned to any classroom.")
            return redirect('teacher_dashboard')
        
        classroom = teacher_enrollment.classroom
        stream = teacher_enrollment.stream

        raw_date = request.GET.get("date")
        if raw_date:
            try:
                selected_date = datetime.strptime(raw_date, "%Y-%m-%d").date()
            except ValueError:
                selected_date = timezone.localdate()
        else:
            selected_date = timezone.localdate()

        # GET ALL STUDENTS in the class (for total counts)
        all_students_qs = StudentProfile.objects.filter(
            enrollments__classroom=classroom,
            enrollments__academic_year=active_year
        ).select_related('user').distinct()
        
        if stream:
            all_students_qs = all_students_qs.filter(enrollments__stream=stream)
        
        # Get ALL student IDs
        all_student_ids = list(all_students_qs.values_list('id', flat=True))
        students_count = len(all_student_ids)
        
        # Get gender counts from ALL students
        male_student_ids = list(all_students_qs.filter(user__gender__iexact='Male').values_list('id', flat=True))
        female_student_ids = list(all_students_qs.filter(user__gender__iexact='Female').values_list('id', flat=True))
        male_count = len(male_student_ids)
        female_count = len(female_student_ids)
        
        # Get attendance for selected date
        attendance_qs = Attendance.objects.filter(
            student_id__in=all_student_ids,
            date=selected_date
        ).select_related('student__user', 'marked_by__user')
        
        # Calculate totals from attendance records
        total_present = attendance_qs.filter(status='present').count()
        total_absent = attendance_qs.filter(status='absent').count()
        total_sick = attendance_qs.filter(status='sick').count()
        total_records = total_present + total_absent + total_sick
        
        # Calculate percentages based on students who have attendance records
        if total_records > 0:
            present_percentage = round((total_present / total_records) * 100, 2)
            absent_percentage = round((total_absent / total_records) * 100, 2)
            sick_percentage = round((total_sick / total_records) * 100, 2)
        else:
            present_percentage = absent_percentage = sick_percentage = 0
        
        # Calculate Male attendance from attendance records
        male_attendance = attendance_qs.filter(student_id__in=male_student_ids)
        male_present = male_attendance.filter(status='present').count()
        male_absent = male_attendance.filter(status='absent').count()
        male_sick = male_attendance.filter(status='sick').count()
        male_total = male_present + male_absent + male_sick
        
        # Calculate Male percentages
        if male_total > 0:
            male_present_pct = round((male_present / male_total) * 100, 2)
            male_absent_pct = round((male_absent / male_total) * 100, 2)
            male_sick_pct = round((male_sick / male_total) * 100, 2)
        else:
            male_present_pct = male_absent_pct = male_sick_pct = 0
        
        # Calculate Female attendance from attendance records
        female_attendance = attendance_qs.filter(student_id__in=female_student_ids)
        female_present = female_attendance.filter(status='present').count()
        female_absent = female_attendance.filter(status='absent').count()
        female_sick = female_attendance.filter(status='sick').count()
        female_total = female_present + female_absent + female_sick
        
        # Calculate Female percentages
        if female_total > 0:
            female_present_pct = round((female_present / female_total) * 100, 2)
            female_absent_pct = round((female_absent / female_total) * 100, 2)
            female_sick_pct = round((female_sick / female_total) * 100, 2)
        else:
            female_present_pct = female_absent_pct = female_sick_pct = 0

        # Pagination for attendance records
        paginator = Paginator(attendance_qs, 25)
        page_number = request.GET.get('page')
        attendance_records = paginator.get_page(page_number)

        context = {
            "classroom": classroom,
            "stream": stream,
            "teacher": teacher,
            "attendance_records": attendance_records,
            "selected_date": selected_date.strftime("%Y-%m-%d"),
            "today": timezone.localdate(),
            "students_count": students_count,
            "total_present": total_present,
            "total_absent": total_absent,
            "total_sick": total_sick,
            "present_percentage": present_percentage,
            "absent_percentage": absent_percentage,
            "sick_percentage": sick_percentage,
            # Male counts
            "male_count": male_count,
            "male_present": male_present,
            "male_absent": male_absent,
            "male_sick": male_sick,
            "male_present_pct": male_present_pct,
            "male_absent_pct": male_absent_pct,
            "male_sick_pct": male_sick_pct,
            # Female counts
            "female_count": female_count,
            "female_present": female_present,
            "female_absent": female_absent,
            "female_sick": female_sick,
            "female_present_pct": female_present_pct,
            "female_absent_pct": female_absent_pct,
            "female_sick_pct": female_sick_pct,
        }

        return render(request, "attendance_app/view_attendance.html", context)
    except Exception as e:
        logger.error(f"Error in view_attendance: {e}")
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect('teacher_dashboard')

@never_cache
@login_required
def edit_attendance(request, pk):
    attendance = get_object_or_404(Attendance, id=pk)
    today = now().date()

    if attendance.date < today and request.method == "POST":
        return JsonResponse({"success": False, "message": "Cannot edit attendance for past days."})

    if request.method == "POST":
        new_status = request.POST.get("status")
        if new_status in ["present", "absent", "sick"]:
            attendance.status = new_status
            attendance.save()
            
            # Hapa tunaiweka kwenye session ya Django Messages kabla ya kurudisha JSON.
            # Ukurasa ukijirefresh kupitia JS, base.html itaikuta na kuonyesha Toast kiotomatiki.
            messages.success(request, f"Attendance updated to {new_status.capitalize()} successfully.")
            
            return JsonResponse({"success": True, "message": f"Attendance updated to {new_status.capitalize()} successfully."})
        return JsonResponse({"success": False, "message": "Invalid status selected."})

    sms_logs = SMSLog.objects.filter(student=attendance.student, timestamp__date=attendance.date)
    can_edit = attendance.date >= today
    return render(request, "attendance_app/edit_attendance_modal.html", {
        "attendance": attendance,
        "sms_logs": sms_logs,
        "can_edit": can_edit
    })


@never_cache
@login_required
def delete_attendance(request, pk):
    if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        try:
            attendance = get_object_or_404(Attendance, pk=pk)
            attendance.delete()
            return JsonResponse({"success": True, "message": "Attendance record deleted successfully."})
        except Exception as e:
            return JsonResponse({"success": False, "message": f"Failed to delete: {str(e)}"})
    return JsonResponse({"success": False, "message": "Invalid request."}, status=400)


@login_required
def attendance_export_pdf(request):
    teacher = get_object_or_404(TeacherProfile, user=request.user)
    classroom, stream = get_teacher_current_assignment(teacher)
    active_year = AcademicYear.objects.filter(is_active=True).first()

    if not classroom:
        messages.error(request, "You are not assigned to any classroom.")
        return redirect('teacher_dashboard')

    raw_date = request.GET.get("date")
    try:
        selected_date = datetime.strptime(raw_date, "%Y-%m-%d").date() if raw_date else now().date()
    except ValueError:
        selected_date = now().date()

    students = get_students_by_teacher_scope(classroom, stream, academic_year=active_year)
    qs = Attendance.objects.filter(student__in=students, date=selected_date).select_related("student", "student__user")

    html = render_to_string("attendance_app/attendance_pdf.html", {
        "records": qs,
        "classroom": classroom,
        "date": selected_date
    })

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename=Attendance_{classroom.name}_{selected_date}.pdf'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse("Error generating PDF", status=500)

    return response


@login_required
def attendance_export_excel(request):
    teacher = get_object_or_404(TeacherProfile, user=request.user)
    classroom, stream = get_teacher_current_assignment(teacher)
    active_year = AcademicYear.objects.filter(is_active=True).first()

    if not classroom:
        messages.error(request, "You are not assigned to any classroom.")
        return redirect('teacher_dashboard')

    raw_date = request.GET.get("date")
    try:
        selected_date = datetime.strptime(raw_date, "%Y-%m-%d").date() if raw_date else now().date()
    except ValueError:
        selected_date = now().date()

    students = get_students_by_teacher_scope(classroom, stream, academic_year=active_year)
    qs = Attendance.objects.filter(student__in=students, date=selected_date).select_related("student", "student__user")

    total = qs.count()
    present = qs.filter(status="present").count()
    absent = qs.filter(status="absent").count()
    sick = qs.filter(status="sick").count()

    def pct(x):
        return round((x / total) * 100, 1) if total else 0

    male_qs = qs.filter(student__user__gender="male")
    female_qs = qs.filter(student__user__gender="female")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value="SMARTPRESENCE – ATTENDANCE REPORT").font = bold
    ws.cell(row=row, column=1).alignment = center
    row += 2

    ws.append(["Class:", classroom.name, "Year:", str(classroom.year), "Date:", selected_date.strftime("%Y-%m-%d")])
    for col in range(1, 7):
        ws.cell(row=row, column=col).font = bold
    row += 2

    ws.append(["SUMMARY"])
    ws.cell(row=row, column=1).font = bold
    row += 1
    ws.append(["Total Students", total])
    ws.append(["Present", f"{present} ({pct(present)}%)"])
    ws.append(["Absent", f"{absent} ({pct(absent)}%)"])
    ws.append(["Sick", f"{sick} ({pct(sick)}%)"])
    row += 2

    ws.append(["GENDER SUMMARY"])
    ws.cell(row=row, column=1).font = bold
    row += 1
    ws.append(["Male", male_qs.count(), f"Present {male_qs.filter(status='present').count()}", f"Absent {male_qs.filter(status='absent').count()}", f"Sick {male_qs.filter(status='sick').count()}"])
    ws.append(["Female", female_qs.count(), f"Present {female_qs.filter(status='present').count()}", f"Absent {female_qs.filter(status='absent').count()}", f"Sick {female_qs.filter(status='sick').count()}"])
    row += 2

    headers = ["No", "Admission No", "Student Name", "Gender", "Status", "Date"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=row, column=col).font = bold
        ws.cell(row=row, column=col).alignment = center
    row += 1

    for i, r in enumerate(qs, start=1):
        ws.append([i, r.student.admission_number, r.student.user.get_full_name(), r.student.user.gender.capitalize() if r.student.user.gender else "", r.status.capitalize(), r.date.strftime("%Y-%m-%d")])

    widths = [6, 18, 28, 12, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename=Attendance_{classroom.name}_{selected_date}.xlsx'
    wb.save(response)
    return response


# ================= SMS LOGS VIEWS =================

@never_cache
@login_required
@user_passes_test(lambda u: u.role in ['admin', 'teacher'])
def sms_logs(request):
    user = request.user

    if user.role == 'admin':
        logs = SMSLog.objects.all().order_by('-timestamp')
    elif user.role == 'teacher':
        try:
            teacher_profile = TeacherProfile.objects.get(user=user)
            active_year = AcademicYear.objects.filter(is_active=True).first()
            teacher_enrollment = Enrollment.objects.filter(
                class_teacher=teacher_profile,
                academic_year=active_year
            ).select_related('classroom', 'stream').first()
            
            if teacher_enrollment and teacher_enrollment.classroom:
                students_in_class = get_students_by_teacher_scope(
                    classroom=teacher_enrollment.classroom,
                    stream=teacher_enrollment.stream,
                    academic_year=active_year
                )
                logs = SMSLog.objects.filter(student__in=students_in_class).order_by('-timestamp')
            else:
                logs = SMSLog.objects.none()
        except TeacherProfile.DoesNotExist:
            logs = SMSLog.objects.none()
    else:
        logs = SMSLog.objects.none()

    paginator = Paginator(logs, 25)
    page_number = request.GET.get('page')
    paginated_logs = paginator.get_page(page_number)

    return render(request, 'attendance_app/sms_logs.html', {'sms_logs': paginated_logs})


@never_cache
@login_required
@user_passes_test(lambda u: u.role in ['admin', 'teacher'])
def delete_sms_log_admin(request, sms_id):
    if request.method == "POST":
        try:
            sms_log = get_object_or_404(SMSLog, id=sms_id)
            sms_log.delete()
            return JsonResponse({'success': True, 'message': 'SMS log deleted successfully'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@never_cache
@login_required
@user_passes_test(lambda u: u.role in ['admin', 'teacher'])
def bulk_delete_sms_logs(request):
    if request.method == "POST":
        try:
            import json
            data = json.loads(request.body)
            ids = data.get('ids', [])
            if not ids:
                return JsonResponse({'success': False, 'message': 'No IDs provided'})
            deleted_count = SMSLog.objects.filter(id__in=ids).delete()[0]
            return JsonResponse({'success': True, 'message': f'{deleted_count} SMS log(s) deleted successfully', 'deleted_count': deleted_count})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@never_cache
@login_required
@user_passes_test(lambda u: u.role == 'teacher')
def teacher_sms_logs(request):
    teacher = get_object_or_404(TeacherProfile, user=request.user)
    active_year = AcademicYear.objects.filter(is_active=True).first()
    
    if not active_year:
        messages.error(request, "No active academic year found.")
        return redirect('teacher_dashboard')
    
    teacher_enrollment = Enrollment.objects.filter(
        class_teacher=teacher,
        academic_year=active_year
    ).select_related('classroom', 'stream').first()
    
    if not teacher_enrollment or not teacher_enrollment.classroom:
        messages.error(request, "You are not assigned to a classroom or stream for the current academic year.")
        return redirect('teacher_dashboard')
    
    classroom = teacher_enrollment.classroom
    stream = teacher_enrollment.stream
    students_in_class = get_students_by_teacher_scope(classroom, stream, academic_year=active_year)
    
    # 1. Pata logs zote za wanafunzi wa darasa husika kwanza
    logs = SMSLog.objects.filter(student__in=students_in_class).select_related('student__user', 'parent__user')
    
    # 2. CHUJA KWA SEARCH QUERY (Kama ipo kwenye GET request)
    search_query = request.GET.get('search', '').strip()
    if search_query:
        logs = logs.filter(
            Q(student__user__first_name__icontains=search_query) |
            Q(student__user__last_name__icontains=search_query) |
            Q(parent__user__first_name__icontains=search_query) |
            Q(parent__user__last_name__icontains=search_query) |
            Q(parent__user__phone_number__icontains=search_query)
        )
    
    # 3. CHUJA KWA STATUS (Kama ipo kwenye GET request)
    selected_status = request.GET.get('status', '').strip()
    if selected_status:
        logs = logs.filter(status=selected_status)
        
    # 4. Panga kwa kupanga kwa kuanzia SMS ya hivi karibuni (Ordering)
    logs = logs.order_by('-timestamp')
    
    # 5. PAGINATION (Kuweka kurasa)
    paginator = Paginator(logs, 25)
    page_number = request.GET.get('page')
    paginated_logs = paginator.get_page(page_number)
    
    # 6. RUDISHA DATA KWENYE TEMPLATE
    return render(request, 'attendance_app/teacher_sms_logs.html', {
        'logs': paginated_logs,
        'teacher': teacher,
        'classroom': classroom,
        'stream': stream,
        'search_query': search_query,        # Inarudisha text iliyotafutwa kwenye input field
        'selected_status': selected_status,  # Inarudisha status iliyochaguliwa kwenye dropdown
    })


from django.http import Http404, JsonResponse
from django.views.decorators.http import require_POST

@never_cache
@login_required
@user_passes_test(lambda u: u.role == 'teacher')
@require_POST
def delete_sms_log(request, sms_id):
    """Delete single SMS log - Returns JSON for toast notifications"""
    try:
        teacher = get_object_or_404(TeacherProfile, user=request.user)
        active_year = AcademicYear.objects.filter(is_active=True).first()
        
        if not active_year:
            return JsonResponse({
                'success': False, 
                'message': '❌ Hakuna mwaka wa masomo ulio hai'
            })
        
        teacher_enrollment = Enrollment.objects.filter(
            class_teacher=teacher,
            academic_year=active_year
        ).select_related('classroom', 'stream').first()
        
        if not teacher_enrollment or not teacher_enrollment.classroom:
            return JsonResponse({
                'success': False, 
                'message': '❌ Hujapewa darasa la kufundisha'
            })
        
        students_in_class = get_students_by_teacher_scope(
            teacher_enrollment.classroom, 
            teacher_enrollment.stream, 
            academic_year=active_year
        )
        
        sms_log = get_object_or_404(SMSLog, id=sms_id, student__in=students_in_class)
        student_name = sms_log.student.user.get_full_name()
        sms_log.delete()
        
        logger.info(f"SMS log {sms_id} deleted by teacher {teacher.user.username}")
        
        return JsonResponse({
            'success': True, 
            'message': f' {student_name} SMS log deleted'
        })
        
    except Http404:
        return JsonResponse({
            'success': False, 
            'message': 'dint found any message'
        })
    except Exception as e:
        logger.error(f"Error deleting SMS log {sms_id}: {e}")
        return JsonResponse({
            'success': False, 
            'message': f'error: {str(e)[:100]}'
        })


@never_cache
@login_required
@user_passes_test(lambda u: u.role == 'teacher')
def bulk_delete_sms_logs_teacher(request):
    """Bulk delete multiple SMS logs - Returns JSON for toast notifications"""
    if request.method != "POST":
        return JsonResponse({
            'success': False, 
            'message': 'method not accepted'
        })
    
    try:
        import json
        data = json.loads(request.body)
        ids = data.get('ids', [])
        
        if not ids:
            return JsonResponse({
                'success': False, 
                'message': 'sms not'
            })
        
        teacher = get_object_or_404(TeacherProfile, user=request.user)
        active_year = AcademicYear.objects.filter(is_active=True).first()
        
        if not active_year:
            return JsonResponse({
                'success': False, 
                'message': 'No active academic year'
            })
        
        teacher_enrollment = Enrollment.objects.filter(
            class_teacher=teacher,
            academic_year=active_year
        ).select_related('classroom', 'stream').first()
        
        if not teacher_enrollment or not teacher_enrollment.classroom:
            return JsonResponse({
                'success': False, 
                'message': 'your not enrolled to class'
            })
        
        students_in_class = get_students_by_teacher_scope(
            teacher_enrollment.classroom, 
            teacher_enrollment.stream, 
            academic_year=active_year
        )
        
        deleted_count = SMSLog.objects.filter(
            id__in=ids, 
            student__in=students_in_class
        ).delete()[0]
        
        logger.info(f"Bulk deleted {deleted_count} SMS logs by teacher {teacher.user.username}")
        
        return JsonResponse({
            'success': True, 
            'message': f' {deleted_count} SMS log(s) success deleted!',
            'deleted_count': deleted_count
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False, 
            'message': 'incorrect data'
        })
    except Exception as e:
        logger.error(f"Error in bulk delete SMS logs: {e}")
        return JsonResponse({
            'success': False, 
            'message': f' Error: {str(e)[:100]}'
        })


@never_cache
@login_required
@user_passes_test(lambda u: u.role == 'teacher')
def resend_sms(request, sms_id):
    """Resend SMS - Returns JSON for toast notifications"""
    try:
        teacher = request.user.teacher_profile
    except TeacherProfile.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'message': 'No teacher profile found'
        })

    active_year = AcademicYear.objects.filter(is_active=True).first()
    if not active_year:
        return JsonResponse({
            'success': False, 
            'message': ' No active academic year'
        })
    
    teacher_enrollment = Enrollment.objects.filter(
        class_teacher=teacher,
        academic_year=active_year
    ).select_related('classroom', 'stream').first()
    
    if not teacher_enrollment or not teacher_enrollment.classroom:
        return JsonResponse({
            'success': False, 
            'message': ' not enrolled'
        })
    
    students_in_class = get_students_by_teacher_scope(
        teacher_enrollment.classroom, 
        teacher_enrollment.stream, 
        academic_year=active_year
    )
    
    sms_log = get_object_or_404(SMSLog, id=sms_id, student__in=students_in_class)
    student_name = sms_log.student.user.get_full_name()

    if not sms_log.parent or not sms_log.parent.user.phone_number:
        return JsonResponse({
            'success': False, 
            'message': f' Namba ya simu ya mzazi wa {student_name} haipo'
        })

    if not sms_log.message or sms_log.message.strip() == "":
        return JsonResponse({
            'success': False, 
            'message': ' Ujumbe wa SMS uko tupu. Haiwezi kutumwa tena.'
        })

    try:
        parent_phone = sms_log.parent.user.phone_number
        sms_sent, response_msg = send_sms(parent_phone, sms_log.message)

        if sms_sent:
            sms_log.status = "sent"
            sms_log.timestamp = timezone.now()
            sms_log.save()
            logger.info(f"SMS resent for {student_name} to {parent_phone}")
            return JsonResponse({
                'success': True, 
                'message': f' SMS imetumwa tena kwa mzazi wa {student_name}!'
            })
        else:
            sms_log.status = "failed"
            sms_log.save()
            
            # Check for balance issue
            if 'salio' in response_msg.lower() or 'balance' in response_msg.lower():
                return JsonResponse({
                    'success': False, 
                    'message': f' {response_msg}'
                })
            else:
                return JsonResponse({
                    'success': False, 
                    'message': f' {response_msg}'
                })
                
    except Exception as e:
        logger.error(f"Resend SMS exception for log {sms_id}: {e}")
        sms_log.status = "failed"
        sms_log.save()
        return JsonResponse({
            'success': False, 
            'message': f' Imeshindikana: {str(e)[:100]}'
        })


# ================= CLASSROOM & ACADEMIC YEAR VIEWS =================

@never_cache
@login_required
@user_passes_test(lambda u: u.role == 'admin')
def manage_classrooms(request):

    all_academic_years = AcademicYear.objects.all().order_by('-year_start')

    selected_year_id = request.GET.get('year')
    selected_academic_year = None

    if selected_year_id:
        try:
            selected_academic_year = AcademicYear.objects.get(id=selected_year_id)
        except AcademicYear.DoesNotExist:
            selected_academic_year = AcademicYear.objects.filter(is_active=True).first()
    else:
        selected_academic_year = AcademicYear.objects.filter(is_active=True).first()

    # ADD CLASSROOM
    if request.method == "POST":

        name = request.POST.get('name', '').strip()
        year_id = request.POST.get('year')

        if not name or not year_id:
            messages.error(
                request,
                "Please provide both classroom name and academic year."
            )
            return redirect('manage_classrooms')

        year_obj = get_object_or_404(AcademicYear, id=year_id)

        # ONLY ACTIVE YEAR
        if not year_obj.is_active:
            messages.error(
                request,
                "You can only add classrooms to the active academic year."
            )
            return redirect('manage_classrooms')

        duplicate = Classroom.objects.filter(
            name__iexact=name,
            year=year_obj
        ).exists()

        if duplicate:
            messages.error(
                request,
                f"Classroom '{name}' already exists for {year_obj}."
            )
            return redirect('manage_classrooms')

        Classroom.objects.create(
            name=name,
            year=year_obj
        )

        messages.success(
            request,
            f"Classroom '{name}' added successfully."
        )

        return redirect('manage_classrooms')

    # FILTER CLASSROOMS
    if selected_academic_year:
        classrooms = Classroom.objects.filter(
            year=selected_academic_year
        ).order_by('name')
    else:
        classrooms = Classroom.objects.all().order_by(
            'year__year_start',
            'name'
        )

    context = {
        'classrooms': classrooms,
        'academic_years': all_academic_years,
        'selected_academic_year': selected_academic_year,
    }

    return render(
        request,
        'attendance_app/manage_classrooms.html',
        context
    )


@never_cache
@login_required
@user_passes_test(lambda u: u.role == 'admin')
def delete_classroom(request, classroom_id):

    classroom = get_object_or_404(Classroom, id=classroom_id)

    # Check students assigned through enrollments
    has_students = classroom.class_enrollments.filter(
        student__isnull=False
    ).exists()

    # Check teachers assigned through enrollments
    has_teachers = classroom.class_enrollments.filter(
        class_teacher__isnull=False
    ).exists()

    if request.method == "POST":

        if has_students or has_teachers:

            parts = []

            if has_students:
                parts.append("students")

            if has_teachers:
                parts.append("teachers")

            parts_str = " and ".join(parts)

            messages.error(
                request,
                f"Cannot delete '{classroom.name}' because it has assigned {parts_str}."
            )

        else:

            # Delete streams first
            classroom.streams.all().delete()

            classroom_name = classroom.name

            classroom.delete()

            messages.success(
                request,
                f"Classroom '{classroom_name}' deleted successfully."
            )

    return redirect('manage_classrooms')


@never_cache
@login_required
@user_passes_test(lambda u: u.role == 'admin')
def edit_classroom(request, classroom_id):

    classroom = get_object_or_404(
        Classroom,
        id=classroom_id
    )

    if request.method == "POST":

        # ONLY ACTIVE YEAR CAN BE EDITED
        if not classroom.year.is_active:

            messages.error(
                request,
                "Only classrooms in the active academic year can be edited."
            )

            return redirect('manage_classrooms')

        new_name = request.POST.get('name', '').strip()
        year_id = request.POST.get('year')

        if not new_name:
            messages.error(
                request,
                "Classroom name cannot be empty."
            )

            return redirect('manage_classrooms')

        year = get_object_or_404(
            AcademicYear,
            id=year_id
        )

        # ONLY ACTIVE YEAR
        if not year.is_active:

            messages.error(
                request,
                "You can only assign classrooms to the active academic year."
            )

            return redirect('manage_classrooms')

        duplicate = Classroom.objects.filter(
            name__iexact=new_name,
            year=year
        ).exclude(id=classroom.id).exists()

        if duplicate:

            messages.error(
                request,
                f"Classroom '{new_name}' already exists."
            )

            return redirect('manage_classrooms')

        classroom.name = new_name
        classroom.year = year
        classroom.save()

        # UPDATE STREAMS
        stream_ids = request.POST.getlist('stream_ids[]')
        stream_names = request.POST.getlist('stream_names[]')

        for stream_id, stream_name in zip(stream_ids, stream_names):

            stream = Stream.objects.filter(
                id=stream_id,
                classroom=classroom
            ).first()

            if stream and stream_name.strip():

                stream.name = stream_name.strip()
                stream.save()

        # DELETE STREAMS
        delete_streams = request.POST.getlist('delete_streams')

        if delete_streams:

            Stream.objects.filter(
                id__in=delete_streams,
                classroom=classroom
            ).delete()

        # ADD NEW STREAM
        new_stream = request.POST.get('new_stream', '').strip()

        if new_stream:

            exists = Stream.objects.filter(
                classroom=classroom,
                name__iexact=new_stream
            ).exists()

            if not exists:

                Stream.objects.create(
                    classroom=classroom,
                    name=new_stream
                )

        messages.success(
            request,
            f"Classroom '{classroom.name}' updated successfully."
        )

        return redirect('manage_classrooms')

    return redirect('manage_classrooms')


@never_cache
@login_required
@user_passes_test(lambda u: u.role == 'admin')
def add_stream(request, class_id):
    classroom = get_object_or_404(Classroom, id=class_id)
    
    if not classroom.year.is_active:
        messages.error(request, "You can only add streams to classrooms in the active academic year.")
        return redirect('manage_classrooms')
    
    if request.method == 'POST':
        stream_name = request.POST.get('stream_name', '').strip().upper()  # Inakubali 'A', 'B', etc.
        
        if not stream_name:
            messages.error(request, "Stream name cannot be empty.")
            return redirect('manage_classrooms')
        
        exists = classroom.streams.filter(name__iexact=stream_name).exists()
        
        if exists:
            messages.error(request, f"Stream '{stream_name}' already exists in classroom '{classroom.name}'.")
        else:
            Stream.objects.create(name=stream_name, classroom=classroom)
            messages.success(request, f"Stream '{stream_name}' added successfully.")
    
    return redirect('manage_classrooms')

@never_cache
@login_required
def academic_years(request):
    auto_lock_expired_academic_year()
    years = AcademicYear.objects.all().order_by('-year_start')
    return render(request, 'attendance_app/academic_years.html', {'years': years})


@never_cache
@login_required
def add_academic_year(request):
    if request.method == 'POST':
        form = AcademicYearForm(request.POST)
        if form.is_valid():
            year_start = form.cleaned_data['year_start']
            year_end = form.cleaned_data['year_end']
            if AcademicYear.objects.filter(year_start=year_start, year_end=year_end).exists():
                messages.error(request, "Academic Year already exists")
            else:
                form.save()
                messages.success(request, "Academic Year added successfully")
                return redirect('academic_years')
    else:
        form = AcademicYearForm()
    return render(request, 'attendance_app/add_academic_year.html', {'form': form})


@never_cache
@login_required
def edit_academic_year(request, id):
    year = get_object_or_404(AcademicYear, id=id)
    if request.method == "POST":
        year_start = request.POST.get('year_start')
        try:
            year_start = int(year_start)
        except ValueError:
            messages.error(request, "Start Year must be a number")
            return redirect('academic_years')
        year.year_start = year_start
        year.year_end = year_start + 1
        year.is_active = 'is_active' in request.POST
        year.save()
        messages.success(request, f"Academic Year {year.year_start}/{year.year_end} updated successfully")
        return redirect('academic_years')
    return redirect('academic_years')


@never_cache
@login_required
def delete_academic_year(request, id):
    year = get_object_or_404(AcademicYear, id=id)
    if request.method == "POST":
        try:
            year.delete()
            messages.success(request, f"Academic Year {year.year_start}/{year.year_end} deleted successfully.")
        except ProtectedError:
            messages.error(request, f"Cannot delete Academic Year {year.year_start}/{year.year_end} because it is in use.")
        return redirect('academic_years')
    return redirect('academic_years')


# ================= ATTENDANCE REPORT VIEWS =================

@never_cache
@login_required
@user_passes_test(lambda u: u.role == 'admin')
def attendance_report(request):
    classroom_id = request.GET.get('classroom')
    attendances = Attendance.objects.select_related('student__user', 'marked_by__user')

    if classroom_id:
        # Fix: Filter through student's enrollment
        attendances = attendances.filter(student__enrollments__classroom_id=classroom_id)

    classrooms = Classroom.objects.all()
    return render(request, 'attendance_app/attendance_report.html', {
        'attendances': attendances,
        'classrooms': classrooms
    })


@never_cache
@login_required
@user_passes_test(lambda u: u.role == 'admin')
def attendance_report_cards(request):
    """
    Show attendance report by cards for each classroom.
    Clicking a card navigates to detailed attendance of that class.
    Supports filtering by academic year.
    """
    # Get all academic years for the dropdown
    all_academic_years = AcademicYear.objects.all().order_by('-year_start')
    
    # Get selected academic year from request
    selected_year_id = request.GET.get('year')
    selected_academic_year = None
    
    if selected_year_id:
        try:
            selected_academic_year = AcademicYear.objects.get(id=selected_year_id)
        except AcademicYear.DoesNotExist:
            selected_academic_year = AcademicYear.objects.filter(is_active=True).first()
    else:
        # Default to active academic year
        selected_academic_year = AcademicYear.objects.filter(is_active=True).first()
    
    # Get classrooms for the selected academic year
    if selected_academic_year:
        classrooms = Classroom.objects.filter(year=selected_academic_year).order_by('name')
        
        # Add attendance summary for each classroom
        for classroom in classrooms:
            # Get students enrolled in this classroom for the selected year
            students = StudentProfile.objects.filter(
                enrollments__classroom=classroom,
                enrollments__academic_year=selected_academic_year
            ).distinct()
            
            classroom.total_students = students.count()
            
            # Get today's attendance summary for this classroom
            today = now().date()
            attendance_today = Attendance.objects.filter(
                student__in=students,
                date=today
            )
            
            classroom.present_today = attendance_today.filter(status='present').count()
            classroom.absent_today = attendance_today.filter(status='absent').count()
            classroom.sick_today = attendance_today.filter(status='sick').count()
    else:
        classrooms = Classroom.objects.none()

    context = {
        "classrooms": classrooms,
        "all_academic_years": all_academic_years,
        "selected_academic_year": selected_academic_year,
    }
    return render(request, "attendance_app/attendance_report_cards.html", context)


@never_cache
@login_required
@user_passes_test(lambda u: u.role == 'admin')
def view_class_attendance(request, classroom_id):
    """
    Detailed attendance for a single class.
    Filterable by date, stream, and academic year.
    """
    try:
        classroom = get_object_or_404(Classroom, id=classroom_id)
        streams = Stream.objects.filter(classroom=classroom)

        selected_stream_id = request.GET.get("stream")
        raw_date = request.GET.get("date")
        
        # NEW: Get selected academic year from request
        selected_year_id = request.GET.get('year')
        
        # Determine which academic year to use
        if selected_year_id:
            try:
                active_year = AcademicYear.objects.get(id=selected_year_id)
            except AcademicYear.DoesNotExist:
                active_year = AcademicYear.objects.filter(is_active=True).first()
        else:
            active_year = AcademicYear.objects.filter(is_active=True).first()

        try:
            selected_date = datetime.strptime(raw_date, "%Y-%m-%d").date() if raw_date else now().date()
        except ValueError:
            selected_date = now().date()

        if not active_year:
            messages.error(request, "No academic year found.")
            return redirect('attendance_report_cards')
        
        # Get all students enrolled in this classroom for that academic year
        students_in_class = StudentProfile.objects.filter(
            enrollments__classroom=classroom,
            enrollments__academic_year=active_year
        ).distinct()
        
        # If stream is selected, filter further
        if selected_stream_id:
            students_in_class = students_in_class.filter(
                enrollments__stream_id=selected_stream_id
            )
        
        # Get attendance records for these students on the selected date
        attendance_qs = Attendance.objects.filter(
            student__in=students_in_class,
            date=selected_date
        ).select_related(
            "student__user", 
            "marked_by__user",
            "enrollment__stream"
        ).order_by("student__user__first_name")
        
        # Get students who don't have attendance records for this date
        students_with_attendance = attendance_qs.values_list('student_id', flat=True)
        students_without_attendance = students_in_class.exclude(id__in=students_with_attendance)
        
        # Get all academic years for the filter dropdown
        all_academic_years = AcademicYear.objects.all().order_by('-year_start')
        
        context = {
            "classroom": classroom,
            "attendance_records": attendance_qs,
            "students_without_attendance": students_without_attendance,
            "streams": streams,
            "selected_date": selected_date.strftime("%Y-%m-%d"),
            "selected_stream_id": int(selected_stream_id) if selected_stream_id else None,
            "active_year": active_year,
            "all_academic_years": all_academic_years,
            "selected_year_id": selected_year_id,
        }

        return render(request, "attendance_app/view_class_attendance.html", context)

    except Exception as e:
        logger.error(f"Error in view_class_attendance for class {classroom_id}: {e}")
        messages.error(request, "Could not load attendance data. Please try again or contact support.")
        return redirect('attendance_report_cards')


# ================= PROFILE & SETTINGS VIEWS =================

@never_cache
@login_required
def teacher_profile_view(request):
    teacher_profile = get_object_or_404(TeacherProfile, user=request.user)
    active_year = AcademicYear.objects.filter(is_active=True).first()
    current_enrollment = teacher_profile.class_enrollments.filter(academic_year=active_year).first()

    if request.method == "POST":
        if "update_profile" in request.POST:
            user = request.user
            user.first_name = request.POST.get("first_name", user.first_name)
            user.last_name = request.POST.get("last_name", user.last_name)
            user.username = request.POST.get("username", user.username)
            user.email = request.POST.get("email", user.email)
            
            phone_number = request.POST.get("phone_number", "").strip()
            if phone_number:
                formatted_phone = format_phone_number(phone_number)
                if not formatted_phone:
                    messages.error(request, "Invalid phone number format! Use format: 0712345678 or +255712345678")
                    return redirect('teacher_profile_view')
                user.phone_number = formatted_phone
            else:
                user.phone_number = ""
            
            user.save()
            messages.success(request, "Profile updated successfully!")
            return redirect('teacher_profile_view')

        elif "change_password" in request.POST:
            # Password change logic (same as before)
            ...

    phone_number_display = request.user.phone_number or ""
    current_classroom = current_enrollment.classroom if current_enrollment else None
    current_stream = current_enrollment.stream if current_enrollment else None
    
    context = {
        "teacher": teacher_profile,
        "current_enrollment": current_enrollment,
        "current_classroom_name": current_classroom.name if current_classroom else "Not Assigned",
        "current_stream_name": current_stream.name if current_stream else "Not Assigned",
        "phone_number_display": phone_number_display,
        "active_year": active_year,
    }
    return render(request, "attendance_app/teacher_profile.html", context)


@never_cache
@login_required
@user_passes_test(lambda u: u.role == 'admin')
def admin_profile(request):

    profile_success = None
    password_success = None

    if request.method == "POST":

        # ================= UPDATE PROFILE =================
        if request.POST.get('update_profile'):
            user = request.user
            user.first_name = request.POST.get(
                "first_name",
                user.first_name
            )

            user.last_name = request.POST.get(
                "last_name",
                user.last_name
            )

            user.username = request.POST.get(
                "username",
                user.username
            )

            user.email = request.POST.get(
                "email",
                user.email
            )

            # PHONE NUMBER
            phone_number = request.POST.get(
                "phone_number",
                ""
            ).strip()

            if phone_number:

                formatted_phone = format_phone_number(phone_number)

                if not formatted_phone:
                    messages.error(
                        request,
                        "Invalid phone number format!"
                    )
                    return redirect('admin_profile')

                user.phone_number = formatted_phone

            else:
                user.phone_number = ""

            # SAVE USER
            user.save()

            profile_success = "Profile updated successfully!"

        # ================= CHANGE PASSWORD =================
        elif request.POST.get('change_password'):

            old_password = request.POST.get("old_password")
            new_password1 = request.POST.get("new_password1")
            new_password2 = request.POST.get("new_password2")

            if request.user.check_password(old_password):

                if new_password1 == new_password2:

                    request.user.set_password(new_password1)
                    request.user.save()

                    update_session_auth_hash(
                        request,
                        request.user
                    )

                    password_success = "Password changed successfully!"

                else:
                    password_success = "New passwords do not match!"

            else:
                password_success = "Old password is incorrect!"

    context = {
        "profile_success": profile_success,
        "password_success": password_success,
    }

    return render(
        request,
        "attendance_app/admin_profile.html",
        context
    )


@never_cache
@login_required
def school_settings(request):
    try:
        school_settings, _ = SchoolSettings.objects.get_or_create(
            id=1,
            defaults={"school_name": "My School"}
        )

        form = SchoolSettingsForm(
            request.POST or None,
            request.FILES or None,
            instance=school_settings
        )

        if request.method == 'POST':
            if form.is_valid():
                try:
                    if 'logo' in request.FILES and school_settings.logo:
                        if default_storage.exists(school_settings.logo.name):
                            default_storage.delete(school_settings.logo.name)
                    form.save()
                    messages.success(request, "School settings updated successfully")
                    return redirect('school_settings')
                except Exception as e:
                    messages.error(request, f"Upload failed: {str(e)}")
            else:
                messages.error(request, "Failed to update school settings")
    except Exception as e:
        messages.error(request, f"System error: {str(e)}")
        form = None
        school_settings = None

    return render(request, 'attendance_app/admin_settings.html', {
        'form': form,
        'school_settings': school_settings,
    })


# ================= STUDENT EXPORT VIEWS =================

@never_cache
@login_required
def export_students_excel(request):
    teacher = get_object_or_404(TeacherProfile, user=request.user)
    classroom, stream = get_teacher_current_assignment(teacher)
    active_year = AcademicYear.objects.filter(is_active=True).first()

    if not classroom:
        messages.error(request, "You are not assigned to any classroom.")
        return redirect('teacher_dashboard')

    students = get_students_by_teacher_scope(classroom, stream, academic_year=active_year).select_related("user")

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="students_{classroom.name}.csv"'

    writer = csv.writer(response)
    writer.writerow(["Admission No", "Name", "Email", "Phone", "Gender", "Classroom"])

    for s in students:
        writer.writerow([
            s.admission_number,
            f"{s.user.first_name} {s.user.last_name}",
            s.user.email,
            s.user.phone_number or '—',
            s.user.gender.capitalize() if s.user.gender else '—',
            classroom.name
        ])

    return response


@never_cache
@login_required
def export_students_pdf(request):
    teacher = get_object_or_404(TeacherProfile, user=request.user)
    classroom, stream = get_teacher_current_assignment(teacher)
    active_year = AcademicYear.objects.filter(is_active=True).first()

    if not classroom:
        messages.error(request, "You are not assigned to any classroom.")
        return redirect('teacher_dashboard')

    students = get_students_by_teacher_scope(classroom, stream, academic_year=active_year).select_related("user")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="students_{classroom.name}.pdf"'

    p = canvas.Canvas(response)
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, 800, f"Students List - {classroom.name}")
    p.setFont("Helvetica", 10)

    y = 770
    p.drawString(50, y, "No")
    p.drawString(80, y, "Admission No")
    p.drawString(180, y, "Name")
    p.drawString(330, y, "Email")
    p.drawString(480, y, "Phone")
    p.drawString(550, y, "Gender")
    y -= 20

    for i, s in enumerate(students, start=1):
        p.drawString(50, y, str(i))
        p.drawString(80, y, s.admission_number)
        p.drawString(180, y, f"{s.user.first_name} {s.user.last_name}")
        p.drawString(330, y, s.user.email)
        p.drawString(480, y, s.user.phone_number or '—')
        p.drawString(550, y, s.user.gender.capitalize() if s.user.gender else '—')
        y -= 20
        if y < 50:
            p.showPage()
            y = 800
            p.setFont("Helvetica", 10)

    p.showPage()
    p.save()
    return response


# ================= PASSWORD RESET VIEWS =================

@never_cache
def forgot_password(request):
    if request.method == "POST":
        try:
            identifier = request.POST.get("identifier", "").strip()
            if not identifier:
                messages.error(request, "Enter email or phone number")
                return redirect("forgot_password")

            formatted_phone = format_phone_number(identifier)
            user = (User.objects.filter(email=identifier).first() or User.objects.filter(phone_number=formatted_phone).first())

            if not user:
                messages.error(request, "User not found")
                return redirect("forgot_password")

            code = random.randint(100000, 999999)
            request.session["reset_code"] = str(code)
            request.session["reset_user"] = user.id
            request.session["reset_time"] = int(time.time())

            if user.phone_number:
                phone = format_phone_number(user.phone_number)
                try:
                    sms_sent = send_sms(phone, f"Namba ya kurejesha password ni {code}")
                    if sms_sent:
                        logger.info(f"Password reset SMS sent to {phone}")
                except Exception as sms_error:
                    logger.error(f"SMS error: {sms_error}")
            return redirect("verify_reset")
        except Exception as e:
            logger.critical(f"Forgot password error: {e}")
            messages.error(request, "Something went wrong. Try again.")
            return redirect("forgot_password")
    return render(request, "attendance_app/forgot_password.html")


def verify_reset(request):
    if request.method == "POST":
        code = request.POST.get("code")
        session_code = request.session.get("reset_code")
        reset_time = request.session.get("reset_time")

        if not session_code or not reset_time:
            messages.error(request, "Session expired")
            return redirect("forgot_password")

        if time.time() - reset_time > 300:
            messages.error(request, "Code expired")
            return redirect("forgot_password")

        if code == session_code:
            return redirect("reset_password")
        messages.error(request, "Invalid code")
    return render(request, "attendance_app/verify_reset.html")


def reset_password(request):
    if request.method == "POST":
        password = request.POST.get("password")
        confirm = request.POST.get("confirm_password")

        if password != confirm:
            messages.error(request, "Passwords do not match")
            return redirect("reset_password")

        user_id = request.session.get("reset_user")
        if not user_id:
            messages.error(request, "Session expired")
            return redirect("forgot_password")

        user = User.objects.get(id=user_id)
        user.set_password(password)
        user.save()
        request.session.flush()
        messages.success(request, "Password reset successful. Please login.")
        return redirect("login")
    return render(request, "attendance_app/reset_password.html")


# ================= HELPER FUNCTIONS FOR SMS =================

def normalize_parent_phone(parent_phone):
    if not parent_phone:
        return None
    parent_phone = re.sub(r"[^\d+]", "", parent_phone)
    if parent_phone.startswith("0") and len(parent_phone) == 10:
        parent_phone = "+255" + parent_phone[1:]
    elif parent_phone.startswith(("6", "7")) and len(parent_phone) == 9:
        parent_phone = "+255" + parent_phone
    elif parent_phone.startswith("255") and len(parent_phone) == 12:
        parent_phone = "+" + parent_phone
    if re.match(r"^\+255[67]\d{8}$", parent_phone):
        return parent_phone
    return None


def send_absent_sms(student, teacher=None):
    """
    Send SMS to all parents of absent student
    Returns: (success_count, failed_count, messages)
    """
    parents = ParentProfile.objects.filter(student=student).select_related('user')
    
    if not parents.exists():
        logger.warning(f"No parent found for student {student.user.get_full_name()}")
        return 0, 1, ["No parent registered for this student"]

    if teacher is None:
        active_year = AcademicYear.objects.filter(is_active=True).first()
        classroom_ids = student.enrollments.filter(
            academic_year=active_year, 
            status='Active'
        ).values_list('classroom_id', flat=True)
        teacher = TeacherProfile.objects.filter(
            class_enrollments__classroom_id__in=classroom_ids,
            class_enrollments__academic_year=active_year
        ).first()

    teacher_phone = teacher.user.phone_number if teacher and teacher.user.phone_number else "N/A"
    teacher_name = teacher.user.get_full_name() if teacher else "Mwalimu"
    
    message_template = (
        f"HABARI MZAZI: Mtoto wako {student.user.get_full_name()} "
        f"hajafika shuleni leo {timezone.now().strftime('%d/%m/%Y')}. "
        f"Tafadhali wasiliana na mwalim {teacher_name} kwa namba {teacher_phone}. Asante."
    )

    success_count = 0
    failed_count = 0
    messages_list = []
    
    for parent in parents:
        parent_phone_raw = parent.user.phone_number
        parent_phone = normalize_parent_phone(parent_phone_raw)

        if not parent_phone:
            failed_count += 1
            msg = f"Invalid phone for {parent.user.get_full_name()}"
            messages_list.append(msg)
            SMSLog.objects.create(
                student=student, 
                parent=parent, 
                message=message_template, 
                status='failed'
            )
            continue

        try:
            success, response_msg = send_sms(parent_phone, message_template)
            status = 'sent' if success else 'failed'
            
            SMSLog.objects.create(
                student=student, 
                parent=parent, 
                message=message_template, 
                status=status
            )
            
            if success:
                success_count += 1
                logger.info(f"SMS sent to {parent_phone} for {student.user.get_full_name()}")
            else:
                failed_count += 1
                messages_list.append(f"{parent.user.get_full_name()}: {response_msg}")
                logger.error(f"SMS failed to {parent_phone}: {response_msg}")
                
        except Exception as e:
            failed_count += 1
            error_msg = str(e)
            messages_list.append(f"{parent.user.get_full_name()}: {error_msg[:50]}")
            logger.error(f"SMS exception for {student.id}: {e}")
            SMSLog.objects.create(
                student=student, 
                parent=parent, 
                message=message_template, 
                status='failed'
            )
    
    return success_count, failed_count, messages_list


# ================= ACADEMIC YEAR PROMOTION VIEWS =================


def get_next_form_name(current_name):
    """
    Reliable promotion: Form I → Form II → Form III → Form IV → Graduate
    Form V → Form VI → Graduate
    """
    if not current_name:
        return None

    name = str(current_name).strip().lower()

    # Promotion mapping - Form IV graduates (does NOT go to Form V)
    mapping = {
        "form i": "Form II",
        "form 1": "Form II",
        "form ii": "Form III",
        "form 2": "Form III",
        "form iii": "Form IV",
        "form 3": "Form IV",
        "form iv": None,      # Form IV GRADUATES - does NOT go to Form V!
        "form 4": None,       # Form IV GRADUATES
        "form four": None,    # Form IV GRADUATES
        "form v": "Form VI",  # Form V goes to Form VI
        "form 5": "Form VI",
        "form vi": None,      # Form VI graduates
        "form 6": None,
    }

    if name in mapping:
        return mapping[name]

    # fallback: extract number
    match = re.search(r'(\d+)', name)
    if match:
        num = int(match.group(1))
        if num == 4 or num >= 6:  # Form IV or Form VI+ graduates
            return None
        return f"Form {num + 1}"

    return None


@transaction.atomic
def generate_academic_year(request):
    """
    FIXED VERSION: Safe academic year generation + promotion system
    - Students: Only ACTIVE students are promoted (Inactive students stay behind)
    - Form IV graduates (does NOT go to Form V)
    - Form VI graduates
    - Teachers: Only ACTIVE teachers are promoted (Inactive teachers stay behind)
    """

    last_year = AcademicYear.objects.order_by('-year_start').first()

    # ================= FIRST YEAR =================
    if not last_year:
        current_year = timezone.now().year

        new_year = AcademicYear.objects.create(
            year_start=current_year,
            is_active=True,
            is_locked=False
        )

        messages.success(request, f"First academic year {new_year} created!")
        return redirect('academic_years')

    new_start = last_year.year_start + 1

    if AcademicYear.objects.filter(year_start=new_start).exists():
        messages.error(request, f"Academic Year {new_start}/{new_start+1} already exists!")
        return redirect('academic_years')

    try:
        # ================= LOCK OLD YEAR SAFELY =================
        old_year = last_year
        old_year.is_active = False
        old_year.is_locked = True
        old_year.save()

        # ================= CREATE NEW YEAR =================
        new_year = AcademicYear.objects.create(
            year_start=new_start,
            is_active=True,
            is_locked=False
        )

        # ================= COPY CLASSROOMS + STREAMS =================
        classroom_map = {}

        for cls in Classroom.objects.filter(year=old_year):
            new_cls = Classroom.objects.create(
                name=cls.name,
                year=new_year
            )
            classroom_map[cls.id] = new_cls

            for st in cls.streams.all():
                Stream.objects.create(
                    name=st.name,
                    classroom=new_cls
                )

        # ================= STUDENT PROMOTION =================
        # Only promote ACTIVE students (Inactive/Promoted/Graduated students are NOT promoted)
        students_promoted = 0
        students_graduated = 0
        students_repeating = 0
        students_failed = 0
        students_skipped_inactive = 0
        students_form4_graduated = 0
        students_form6_graduated = 0

        enrollments = Enrollment.objects.filter(
            academic_year=old_year,
            status='Active',  # ONLY active students are promoted!
            student__isnull=False
        ).select_related('student', 'classroom', 'stream')

        for enr in enrollments:
            student = enr.student
            old_class = enr.classroom

            if not student or not old_class:
                students_failed += 1
                continue

            # Check if this is Form IV (should graduate)
            is_form_four = old_class.name.lower() in ['form iv', 'form 4', 'form four']
            
            # Check if this is Form VI (should graduate)
            is_form_six = old_class.name.lower() in ['form vi', 'form 6', 'form six']
            
            next_class = get_next_form_name(old_class.name)

            # ===== FORM IV GRADUATION =====
            if is_form_four:
                enr.status = 'Graduated'
                enr.save()
                students_graduated += 1
                students_form4_graduated += 1
                continue
            
            # ===== FORM VI GRADUATION =====
            if is_form_six or (next_class is None and not is_form_four):
                enr.status = 'Graduated'
                enr.save()
                students_graduated += 1
                students_form6_graduated += 1
                continue

            # ===== FIND NEW CLASSROOM for Form I, II, III, V =====
            new_class = Classroom.objects.filter(
                name=next_class,
                year=new_year
            ).first()

            # ===== IF NOT FOUND → REPEAT SAME CLASS =====
            if not new_class:
                new_class = classroom_map.get(old_class.id)

                if not new_class:
                    students_failed += 1
                    continue

                students_repeating += 1
            else:
                students_promoted += 1

            # ===== STREAM MATCH =====
            new_stream = None
            if enr.stream:
                new_stream = Stream.objects.filter(
                    name=enr.stream.name,
                    classroom=new_class
                ).first()

            # ===== CREATE NEW ENROLLMENT FOR NEW YEAR =====
            Enrollment.objects.get_or_create(
                student=student,
                academic_year=new_year,
                defaults={
                    'classroom': new_class,
                    'stream': new_stream,
                    'status': 'Active'  # Students start as Active in new year
                }
            )

            # Mark old enrollment as Promoted
            enr.status = 'Promoted'
            enr.save()

        # Count inactive students that were skipped
        students_skipped_inactive = Enrollment.objects.filter(
            academic_year=old_year,
            status='Inactive',
            student__isnull=False
        ).count()

        # ================= TEACHER PROMOTION =================
        teachers_promoted = 0
        teachers_failed = 0
        teachers_skipped_inactive = 0

        teacher_assignments = Enrollment.objects.filter(
            academic_year=old_year,
            class_teacher__isnull=False,
            status='Active'
        ).select_related('class_teacher', 'classroom', 'stream').distinct()

        for assign in teacher_assignments:
            teacher = assign.class_teacher
            old_class = assign.classroom

            if not teacher or not old_class:
                teachers_failed += 1
                continue

            next_class = get_next_form_name(old_class.name)

            if next_class is None:
                next_class = "Form I"

            new_class = Classroom.objects.filter(
                name=next_class,
                year=new_year
            ).first()

            if not new_class:
                teachers_failed += 1
                continue

            new_stream = None
            if assign.stream:
                new_stream = Stream.objects.filter(
                    name=assign.stream.name,
                    classroom=new_class
                ).first()

            Enrollment.objects.update_or_create(
                class_teacher=teacher,
                academic_year=new_year,
                defaults={
                    'classroom': new_class,
                    'stream': new_stream,
                    'student': None,
                    'status': 'Active'
                }
            )

            teachers_promoted += 1

        # Count inactive teachers that were skipped
        teachers_skipped_inactive = Enrollment.objects.filter(
            academic_year=old_year,
            class_teacher__isnull=False,
            status='Inactive'
        ).count()

        # ================= SUCCESS MESSAGE =================
        messages.success(request,
            f"""Academic Year {new_year} created successfully!

STUDENTS:
   Promoted: {students_promoted}
   Repeating: {students_repeating}
   Graduated: {students_graduated}
      Form IV Graduates: {students_form4_graduated}
      Form VI Graduates: {students_form6_graduated}
   Skipped (Inactive): {students_skipped_inactive}
   Failed: {students_failed}

TEACHERS:
   Promoted (Active): {teachers_promoted}
   Skipped (Inactive): {teachers_skipped_inactive}
   Failed: {teachers_failed}

 Old year {old_year.year_start}/{old_year.year_end} has been locked.
 New year {new_year.year_start}/{new_year.year_end} is now ACTIVE.
"""
        )

        return redirect('academic_years')

    except Exception as e:
        logger.error(f"Error generating academic year: {str(e)}")
        messages.error(request, f" Error generating academic year: {str(e)}")
        return redirect('academic_years')
        
    
@never_cache
@login_required
def academic_year_summary(request):
    """OPTIMIZED - Single query version"""
    try:
        years = AcademicYear.objects.all().order_by('-year_start')
        active_year_id = request.GET.get('year')
        active_year = None
        summary_data = []

        if active_year_id:
            try:
                active_year = AcademicYear.objects.get(id=active_year_id)
            except AcademicYear.DoesNotExist:
                active_year = AcademicYear.objects.filter(is_active=True).first()
        else:
            active_year = AcademicYear.objects.filter(is_active=True).first()

        if active_year:
            # OPTIMIZED: Single query with all relations
            classrooms = Classroom.objects.filter(year=active_year).prefetch_related(
                'streams',
                'class_enrollments__student__user',
                'class_enrollments__class_teacher__user'
            )
            
            # Get all enrollments in ONE query
            enrollments = Enrollment.objects.filter(
                academic_year=active_year
            ).select_related(
                'student__user',
                'classroom',
                'class_teacher__user'
            )
            
            # Group by classroom
            classroom_data = {}
            for enrollment in enrollments:
                if enrollment.classroom:
                    cid = enrollment.classroom.id
                    if cid not in classroom_data:
                        classroom_data[cid] = {
                            'classroom': enrollment.classroom,
                            'student_ids': set(),
                            'teachers': set()
                        }
                    if enrollment.student:
                        classroom_data[cid]['student_ids'].add(enrollment.student.id)
                    if enrollment.class_teacher:
                        classroom_data[cid]['teachers'].add(enrollment.class_teacher)
            
            # Get attendance in ONE query
            all_student_ids = set()
            for data in classroom_data.values():
                all_student_ids.update(data['student_ids'])
            
            from django.db.models import Count, Q
            attendance_stats = Attendance.objects.filter(
                student_id__in=all_student_ids
            ).values('student_id', 'status').annotate(total=Count('id'))
            
            # Build summary
            for data in classroom_data.values():
                classroom = data['classroom']
                students_count = len(data['student_ids'])
                
                total_present = total_absent = total_sick = 0
                for stat in attendance_stats:
                    if stat['student_id'] in data['student_ids']:
                        if stat['status'] == 'present':
                            total_present += stat['total']
                        elif stat['status'] == 'absent':
                            total_absent += stat['total']
                        elif stat['status'] == 'sick':
                            total_sick += stat['total']
                
                total_records = total_present + total_absent + total_sick
                
                if students_count > 0 and total_records > 0:
                    present_percentage = round((total_present / total_records) * 100, 2)
                    absent_percentage = round((total_absent / total_records) * 100, 2)
                    sick_percentage = round((total_sick / total_records) * 100, 2)
                else:
                    present_percentage = absent_percentage = sick_percentage = 0
                
                summary_data.append({
                    'classroom': classroom,
                    'students_count': students_count,
                    'teachers': list(data['teachers']),
                    'total_present': total_present,
                    'total_absent': total_absent,
                    'total_sick': total_sick,
                    'present_percentage': present_percentage,
                    'absent_percentage': absent_percentage,
                    'sick_percentage': sick_percentage,
                })

        context = {
            'years': years,
            'active_year': active_year,
            'summary_data': summary_data,
            'terms': [('TERM1', 'Term 1'), ('TERM2', 'Term 2'), ('TERM3', 'Term 3')],
            'selected_term': request.GET.get('term'),
        }
        return render(request, 'attendance_app/academic_year_summary.html', context)
    except Exception as e:
        logger.critical(f"Error in academic_year_summary: {e}")
        messages.error(request, "A critical error occurred.")
        return redirect('academic_years')

# ================= OTHER UTILITY VIEWS =================

@never_cache
@login_required
def student_profile_modal(request, pk):
    student = get_object_or_404(StudentProfile, id=pk)
    return render(request, "attendance_app/student_profile_modal.html", {"student": student})